import pandas as pd
from openpyxl import load_workbook

def read_excel_with_merged_cells(filepath, sheet_name=0):
  # Load workbook and worksheet
  wb = load_workbook(filename=filepath, data_only=True)
  ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[
    sheet_name]

  # Build a matrix with the values
  max_row = ws.max_row
  max_col = ws.max_column
  data = [[None for _ in range(max_col)] for _ in range(max_row)]

  # Fill values from merged cells
  for merged_range in ws.merged_cells.ranges:
    min_row, max_row_range = merged_range.min_row, merged_range.max_row
    min_col, max_col_range = merged_range.min_col, merged_range.max_col
    top_left_value = ws.cell(row=min_row, column=min_col).value
    for row in range(min_row, max_row_range + 1):
      for col in range(min_col, max_col_range + 1):
        data[row - 1][col - 1] = top_left_value

  # Fill in remaining cells (non-merged or not already filled)
  for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1,
                          max_col=ws.max_column):
    for cell in row:
      if data[cell.row - 1][cell.column - 1] is None:
        data[cell.row - 1][cell.column - 1] = cell.value

  # Convert to DataFrame
  df = pd.DataFrame(data)
  return df
