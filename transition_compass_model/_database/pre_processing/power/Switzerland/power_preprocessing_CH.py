import deepl
import pandas as pd
import requests
import os
import pickle
import zipfile
import io
from model.common.io_database import database_to_dm
from model.common.data_matrix_class import DataMatrix
from model.common.auxiliary_functions import create_years_list, my_pickle_dump
import numpy as np
from openpyxl import load_workbook


# Initialize the Deepl Translator
deepl_api_key = '9ecffb3f-5386-4254-a099-8bfc47167661:fx'
translator = deepl.Translator(deepl_api_key)

def df_fso_excel_to_dm(df, header_row, names_dict, var_name, unit, num_cat, keep_first=False, country='Switzerland'):
    # Federal statistical office df from excel to dm
    # Change headers
    new_header = df.iloc[header_row]
    new_header.values[0] = 'Variables'
    df.columns = new_header
    df = df[header_row+1:].copy()
    # Remove nans and empty columns/rows
    if np.nan in df.columns:
        df.drop(columns=np.nan, inplace=True)
    df.set_index('Variables', inplace=True)
    df.dropna(axis=0, how='all', inplace=True)
    df.dropna(axis=1, how='all', inplace=True)
    # Filter rows that contain at least one number (integer or float)
    df = df[df.apply(lambda row: row.map(pd.api.types.is_number), axis=1).any(axis=1)]
    df_clean = df.loc[:, df.apply(lambda col: col.map(pd.api.types.is_number)).any(axis=0)].copy()
    # Extract only the data we are interested in:
    df_filter = df_clean.loc[names_dict.keys()].copy()
    df_filter = df_filter.apply(lambda col: pd.to_numeric(col, errors='coerce'))
    #df_filter = df_filter.applymap(lambda x: pd.to_numeric(x, errors='coerce'))
    df_filter.reset_index(inplace=True)
    # Keep only first 10 caracters
    df_filter['Variables'] = df_filter['Variables'].replace(names_dict)
    if keep_first:
        df_filter = df_filter.drop_duplicates(subset=['Variables'], keep='first')
    df_filter = df_filter.groupby(['Variables']).sum()
    df_filter.reset_index(inplace=True)

    # Pivot the dataframe
    df_filter['Country'] = country
    df_T = pd.melt(df_filter, id_vars=['Variables', 'Country'], var_name='Years', value_name='values')
    df_pivot = df_T.pivot_table(index=['Country', 'Years'], columns=['Variables'], values='values', aggfunc='sum')
    df_pivot = df_pivot.add_suffix('[' + unit + ']')
    df_pivot = df_pivot.add_prefix(var_name + '_')
    df_pivot.reset_index(inplace=True)

    # Drop non numeric values in Years col
    df_pivot['Years'] = pd.to_numeric(df_pivot['Years'], errors='coerce')
    df_pivot = df_pivot.dropna(subset=['Years'])

    dm = DataMatrix.create_from_df(df_pivot, num_cat=num_cat)
    return dm


def translate_text(text):

    translation = translator.translate_text(text, target_lang='EN-GB')
    return translation.text

def create_ots_years_list(years_setting):
    startyear: int = years_setting[0]  # Start year is argument [0], i.e., 1990
    baseyear: int = years_setting[1]  # Base/Reference year is argument [1], i.e., 2015
    lastyear: int = years_setting[2]  # End/Last year is argument [2], i.e., 2050
    step_fts = years_setting[3]  # Timestep for scenario is argument [3], i.e., 5 years
    years_ots = list(np.linspace(start=startyear, stop=baseyear, num=(baseyear - startyear) + 1).astype(int).astype(str))
    return years_ots


def extract_energy_data(file_url, local_filename, baseyear, years_ots, outfile_dm):

    # If the routine had already run and the dm was created, it skips everything and it just loads the dm
    # this is mainly done to avoid calling deepl on repeat which only allows a limited number of calls
    if not os.path.exists(outfile_dm):
        response = requests.get(file_url, stream=True)
        if not os.path.exists(local_filename):
            # Check if the request was successful
            if response.status_code == 200:
                with open(local_filename, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                print(f"File downloaded successfully as {local_filename}")
            else:
                print(f"Error: {response.status_code}, {response.text}")
        else:
            print(f'File {local_filename} already exists. If you want to download again delete the file')

        df = pd.read_csv(local_filename)

        # Clean file format
        # Add """ before the first comma
        df['Jahr'] = df['Jahr'].str.replace(',', '",', 1)
        df = df[['Jahr']]
        # Split the cell into different columns based on ", " separator
        df[['timescale', 'Variable', 'Description', 'value']] = df['Jahr'].str.split('",', expand=True)
        # Drop the original column
        df.drop(columns=['Jahr'], inplace=True)
        for col in df.columns:
            df[col] = df[col].str.lstrip('"')

        # Use deepl to translate variables from de to en
        variables_de = list(set(df['Variable']))
        variables_en = [translate_text(var) for var in variables_de]
        var_dict = dict(zip(variables_de, variables_en))
        # Use deepl to translate description from de to en
        description_de = list(set(df['Description']))
        description_en = [translate_text(var) for var in description_de]
        cat_dict = dict(zip(description_de, description_en))

        df['Use'] = df['Variable'].map(var_dict)
        df['Energy sources'] = df['Description'].map(cat_dict)

        df.drop(['Variable', 'Description'], axis=1, inplace=True)

        # Create database format
        lever = 'energy_tmp'  # temporary lever name
        df['variables'] = df['Use'] + '_' + df['Energy sources'] + '[TJ]'
        df.drop(['Use', 'Energy sources'], axis=1, inplace=True)
        df['level'] = 0
        df['lever'] = 'energy_tmp'
        df['string-pivot'] = 'none'
        df['geoscale'] = 'Switzerland'

        # Replace empty strings with NaN
        df['value'].replace('', np.nan, inplace=True)

        dict_ots, dict_fts = database_to_dm(df, lever, num_cat=1, baseyear=baseyear, years=years_ots, level='all')

        dm = dict_ots[lever]

        with open(outfile_dm, 'wb') as handle:
            pickle.dump(dm, handle, protocol=pickle.HIGHEST_PROTOCOL)

    else:
        with open(outfile_dm, 'rb') as handle:
            dm = pickle.load(handle)

    return dm

def extract_nexuse_capacity_data(file):
    dm = None
    for sheet_yr in ['2020', '2030', '2040', '2050']:
        df_yr = pd.read_excel(file, sheet_name=sheet_yr)

        df_yr = df_yr.loc[df_yr['Country'] == 'CH']
        df_yr = df_yr[
            ['idGen', 'GenName', 'Technology', 'GenEffic', 'CO2Rate', 'Pmax', 'Pmin', 'StartYr', 'EndYr', 'Emax',
             'SubRegion']]

        df_eff_CO2 = df_yr.groupby(['Technology'])[['GenEffic', 'CO2Rate']].mean()

        years_all = years_ots + years_fts
        df = None
        for y in years_all:
            df_all = df_yr[(df_yr['StartYr'] <= y) & (df_yr['EndYr'] >= y)].copy()
            df_all['Years'] = y
            if df is None:
                df = df_all
            else:
                df = pd.concat([df, df_all], axis=0)

        df_P_E = df.groupby(['Technology', 'SubRegion', 'Years'])[['Pmax', 'Pmin', 'Emax']].sum()

        df_P_E.reset_index(inplace=True)
        df_P_E.rename({'SubRegion': 'Country'}, axis=1, inplace=True)

        df_T = df_P_E.pivot(index=['Country', 'Years'], columns='Technology',
                            values=['Pmax', 'Pmin', 'Emax'])
        #df_T.columns = df_T.columns.swaplevel(0, 1)
        df_T.columns = ['_'.join(col).strip() for col in df_T.columns.values]

        df_T.columns = ['pow_capacity-' + col for col in df_T.columns.values]
        cols = []
        for col in df_T.columns.values:
            if 'Pmax' or 'Pmin' in col:
                cols.append(col+'[MW]')
            elif 'Emax' in col:
                cols.append(col + '[MWh]')
        df_T.columns = cols
        df_T.reset_index(inplace=True)

        # Make sure you have all years for all countries
        # Create a DataFrame with all combinations of countries and years
        countries = df_T['Country'].unique()
        complete_index = pd.MultiIndex.from_product([countries, years_all], names=['Country', 'Years'])
        complete_df = pd.DataFrame(index=complete_index).reset_index()

        # Merge with the original DataFrame
        df_T = pd.merge(complete_df, df_T, on=['Country', 'Years'], how='left')

        if dm is None:
            dm = DataMatrix.create_from_df(df_T, num_cat=1)
        else:
            dm_yr = DataMatrix.create_from_df(df_T, num_cat=1)
            extra_cntr = list(set(dm.col_labels['Country']) - set(dm_yr.col_labels['Country']))
            if len(extra_cntr) > 0:
                dm_yr.add(0, dim='Country', col_label=extra_cntr, dummy=True)
            extra_cntr = list(set(dm_yr.col_labels['Country']) - set(dm.col_labels['Country']))
            if len(extra_cntr) > 0:
                dm.add(0, dim='Country', col_label=extra_cntr, dummy=True)
            extra_tech = list(set(dm_yr.col_labels['Categories1']) - set(dm.col_labels['Categories1']))
            if len(extra_tech) > 0:
                dm.add(0, dim='Categories1', col_label=extra_tech, dummy=True)
            extra_tech = list(set(dm.col_labels['Categories1']) - set(dm_yr.col_labels['Categories1']))
            if len(extra_tech) > 0:
                dm_yr.add(0, dim='Categories1', col_label=extra_tech, dummy=True)
            dm_yr.sort('Country')
            dm_yr.sort('Categories1')
            dm.sort('Country')
            dm.sort('Categories1')
            dm.array = np.fmax(dm.array, dm_yr.array)
    dm.sort('Years')
    dm_CH = dm.groupby({'Switzerland': '.*'}, dim='Country', regex=True, inplace=False)
    dm.append(dm_CH, dim='Country')
    return dm, df_eff_CO2


def extract_renewable_capacity_data(file_url, local_filename):

    if not os.path.exists(local_filename):
        response = requests.get(file_url, stream=True)
        # Check if the request was successful
        if response.status_code == 200:
            with open(local_filename, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            print(f"File downloaded successfully as {local_filename}")
        else:
            print(f"Error: {response.status_code}, {response.text}")
    else:
        print(f'File {local_filename} already exists. If you want to download again delete the file')

    df = pd.read_excel(local_filename, sheet_name="Anhang B")
    # Set the new header
    df.columns = df.iloc[1]
    df = df.loc[[48, 231]].copy()
    df['Technologie'] = df['Technologie'].replace({'Photovoltaikanl. (Netz+Insel)': 'pow_capacity-Pmax_PV-roof[MW]',
                                                   'Windenergieanlagen': 'pow_capacity-Pmax_WindOn[MW]'})
    # Keep "Technologie" column and Years columns
    cols_to_keep = []
    for col in df.columns:
        if isinstance(col, int) or col == 'Technologie':
            cols_to_keep.append(col)
    df = df[cols_to_keep]

    df_melted = df.melt(id_vars=["Technologie"], var_name="Years", value_name="Value")
    df_pivoted = df_melted.pivot(index="Years", columns="Technologie", values="Value").reset_index()
    df_pivoted['Country'] = 'Switzerland'

    dm = DataMatrix.create_from_df(df_pivoted, num_cat=1)

    return dm


def extract_production_data(file_url, local_filename):
    if not os.path.exists(local_filename):
        response = requests.get(file_url, stream=True)
        # Check if the request was successful
        if response.status_code == 200:
            with open(local_filename, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            print(f"File downloaded successfully as {local_filename}")
        else:
            print(f"Error: {response.status_code}, {response.text}")
    else:
        print(f'File {local_filename} already exists. If you want to download again delete the file')

    df = pd.read_excel(local_filename, sheet_name="T24")
    combined_headers = []
    for col1, col2, col3 in zip(df.iloc[4], df.iloc[5], df.iloc[6]):
        combined_headers.append(str(col1) + '-' + str(col2) + '[' + str(col3) + ']')
    # Set the new header
    df.columns = combined_headers
    df = df[7:].copy()

    def is_valid_number(val):
        return isinstance(val, (int, float)) and not pd.isna(val)

    # Apply the function to filter out rows with no valid numeric values
    df = df[df.apply(lambda row: row.map(is_valid_number).any(), axis=1)]
    # Apply similarly for columns if needed
    df = df.loc[:, df.apply(lambda col: col.map(is_valid_number).any())]

    df.rename({'Année-nan[nan]': 'Years'}, axis=1, inplace=True)
    df['Country'] = 'Switzerland'
    df.set_index(['Country', 'Years'], inplace=True)
    # df.columns = ['pow_production_'+col for col in df.columns]
    df = df[df.columns.drop(list(df.filter(regex='[%]')))]
    df.columns = [str.replace(col, 'nan-', '') for col in df.columns]
    df.columns = [str.replace(col, 'nan', '') for col in df.columns]
    df.replace('-', 0, inplace=True)
    df.reset_index(inplace=True)
    df = df.drop('Total[GWh]', axis=1)
    dm = DataMatrix.create_from_df(df, num_cat=0)
    cols = ["Centrales hydrauliques-Centrales au fil de l'eau",
            'Centrales nucléaires-',
            'Centrales thermiques class. et centrales chaleur-force1-Total',
            'Centrales à accumulation',
            'Energies renouvelables diverses3-Chauffages au bois et en partie au bois',
            'Eoliennes', 'Installations au biogaz', 'Installations photo-voltaïques', "Pompage d'accumu-lation-",
            'Production nationale (brute)-', 'Production nette (pompage déduit)-', 'dont\nnon renouvelable',
            'dont renouvelable 2']

    dm_out = dm.groupby({'pow_production_RoR': ".*au fil de l'eau.*", 'pow_production_Nuclear': '.*nucléaire.*',
                         'pow_production_Oil-Gas-Waste': '.*class.*|.*biogaz.*', 'pow_production_Dam-gross': '.*accumulation.*',
                         'pow_production_WindOn': '.*Eoliennes.*',
                         'pow_production_PV-roof': '.*photo.*', 'pow_production_Pump-Open': '.*Pompage.*',
                         'pow_production_Waste': '.*dont renouvelable.*'},
                        regex=True, dim='Variables', inplace=False)
    dm_out.deepen()
    dm_out.operation('Dam-gross', '-', 'Pump-Open', out_col='Dam', dim='Categories1')
    dm_out.drop(dim='Categories1', col_label=['Dam-gross'])
    dm_out.operation('Oil-Gas-Waste', '-', 'Waste', out_col='Oil-Gas', dim='Categories1')
    dm_out.drop(dim='Categories1', col_label='Oil-Gas-Waste')
    dm_out.filter({'Years': years_ots}, inplace=True)
    dm_out.change_unit('pow_production', 1000, 'GWh', 'MWh')
    return dm_out



def read_excel_with_merged_cells(filepath, sheet_name=0):
    # Load workbook and worksheet
    wb = load_workbook(filename=filepath, data_only=True)
    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]

    # Build a matrix with the values
    max_row = ws.max_row
    max_col = ws.max_column
    data = [[None for _ in range(max_col)] for _ in range(max_row)]

    # Fill values from merged cells
    for merged_range in ws.merged_cells.ranges:
        min_row, max_row_range = merged_range.min_row, merged_range.max_row
        min_col, max_col_range = merged_range.min_col, merged_range.max_col
        top_left_value = ws.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row_range + 1):
            for col in range(min_col, max_col_range + 1):
                data[row - 1][col - 1] = top_left_value

    # Fill in remaining cells (non-merged or not already filled)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if data[cell.row - 1][cell.column - 1] is None:
                data[cell.row - 1][cell.column - 1] = cell.value

    # Convert to DataFrame
    df = pd.DataFrame(data)
    return df


def extract_importexport_data(file_url, local_filename, sheet_name, var_name, mapping):
    if not os.path.exists(local_filename):
        response = requests.get(file_url, stream=True)
        # Check if the request was successful
        if response.status_code == 200:
            with open(local_filename, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            print(f"File downloaded successfully as {local_filename}")
        else:
            print(f"Error: {response.status_code}, {response.text}")
    else:
        print(f'File {local_filename} already exists. If you want to download again delete the file')

   # df = pd.read_excel(local_filename, sheet_name="T06")

    df = read_excel_with_merged_cells(local_filename, sheet_name)
    combined_headers = []
    for col1, col2 in zip(df.iloc[4], df.iloc[5]):
        combined_headers.append(str(col1) + '[' + str(col2) + ']')
    # Set the new header
    df.columns = combined_headers
    df = df[6:].copy()

    def is_valid_number(val):
        return isinstance(val, (int, float)) and not pd.isna(val)

    # Apply the function to filter out rows with no valid numeric values
    df = df[df.apply(lambda row: row.map(is_valid_number).any(), axis=1)]
    # Apply similarly for columns if needed
    df = df.loc[:, df.apply(lambda col: col.map(is_valid_number).any())]

    df.rename({'Année[Année]': 'Years'}, axis=1, inplace=True)
    df['Country'] = 'Switzerland'
    df.set_index(['Country', 'Years'], inplace=True)
    # df.columns = ['pow_production_'+col for col in df.columns]
    df.replace('-', 0, inplace=True)
    df.replace('None', 0, inplace=True)
    df.reset_index(inplace=True)
    df = df.drop('Total[TJ]', axis=1)
    col_to_keep = df.columns[df.columns.str.contains("TJ", case=False)].tolist()
    filtered_df = df[['Country', 'Years']+col_to_keep].copy()
    dm = DataMatrix.create_from_df(filtered_df, num_cat=0)
    for key in list(mapping.keys()):
        mapping[var_name + '_' + key] = mapping.pop(key)
    dm_out = dm.groupby(mapping, regex=True, dim='Variables', inplace=False)
    dm_out.deepen()
    dm_out.filter({'Years': years_ots}, inplace=True)
    dm_out.change_unit(var_name, 3.6*1e-3, 'TJ', 'MWh', operator='/')

    return dm_out


def compute_capacity_factor(dm_capacity, dm_production):

    dm_capacity_ots = dm_capacity.filter({'Years': years_ots, 'Country': ['Switzerland']}, inplace=False)
    dm_capacity_ots.groupby({'Oil-Gas': 'Oil|Gas.*'}, regex=True, dim='Categories1', inplace=True)
    missing_cat = set(dm_capacity_ots.col_labels['Categories1']) - set(dm_production.col_labels['Categories1'])
    dm_production.add(0, dummy=True, dim='Categories1', col_label=missing_cat)
    dm_production.append(dm_capacity_ots, dim='Variables')

    # Determine the capacity factor
    # capacity-factor = Production / ( Capacity * 24 * 365)
    # Production comes from Statistical Office and Capacity from Nexus-E
    arr_tmp = dm_production[:, :, 'pow_capacity-Pmax', :] * 24 * 365
    dm_production.add(arr_tmp, dim='Variables', col_label='pow_capacity-E', unit='MWh')
    dm_production.operation('pow_production', '/', 'pow_capacity-E', out_col='pow_capacity-factor', unit='%')

    # Handle
    dm_cap_factor = dm_production.filter({'Variables': ['pow_capacity-factor']})
    arr = dm_cap_factor.array
    arr[np.isinf(arr)] = np.nan
    dm_cap_factor.array = arr
    mean_val = np.nanmean(dm_cap_factor.array, axis=1)
    arr = dm_cap_factor.array[0, :, 0, :]
    nan_indices = np.where(np.isnan(arr))
    arr[nan_indices] = np.take(mean_val, nan_indices[1])
    dm_cap_factor.array[0, :, 0, :] = arr
    dm_production.filter({'Variables': ['pow_production']}, inplace=True)

    return dm_cap_factor, dm_production


def extract_hydro_capacity_at_year(df, yr):
    df.rename({'ZE-Kanton': 'Country'}, axis=1, inplace=True)
    # Keep only active power-plants
    df = df.loc[df['ZE-Status'] == 'im Normalbetrieb']
    # Filter useful variables
    df_P = df[['Country', 'Inst. Pumpenleistung']].copy()
    df = df.loc[df['WKA-Typ'] != 'U']  # Pump capacity only
    df = df[['WKA-Typ', 'Country', 'Max. Leistung ab Generator']]  # RoR and Dam
    df = df.loc[df['WKA-Typ'] != 'P']

    # Group Pump capacity by canton and add years column
    df_P = df_P.groupby(['Country']).sum()
    df_P['Years'] = yr
    df_P.rename({'Inst. Pumpenleistung': 'pow_capacity-Pmax_Pump-Open[MW]'}, axis=1, inplace=True)
    df_P.reset_index(inplace=True)
    dm_P = DataMatrix.create_from_df(df_P, num_cat=1)

    # Group capacity by canton and type
    df = df.groupby(['Country', 'WKA-Typ']).sum()
    df.reset_index(inplace=True)
    df['Years'] = yr
    df['WKA-Typ'] = df['WKA-Typ'].str.replace('L', 'pow_capacity-Pmax_RoR[MW]', 1)
    df['WKA-Typ'] = df['WKA-Typ'].str.replace('S', 'pow_capacity-Pmax_Dam[MW]', 1)

    # Dam RoR
    df.rename({'Max. Leistung ab Generator': 'Pmax'}, axis=1, inplace=True)
    df = df.pivot_table(index=['Country', 'Years'], columns=['WKA-Typ'], values='Pmax', aggfunc='sum')
    df.reset_index(inplace=True)
    dm = DataMatrix.create_from_df(df, num_cat=1)

    dm.append(dm_P, dim='Categories1')

    return dm

def extract_old_hydro_capacity_data(url_dict):

    dm_all = None
    for yr in url_dict.keys():
        local_filename = url_dict[yr]['local_filename']
        file_url = url_dict[yr]['file_url']
        if not os.path.exists(local_filename):
            response = requests.get(file_url, stream=True)
            # Check if the request was successful
            if response.status_code == 200:
                with open(local_filename, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                print(f"File downloaded successfully as {local_filename}")
            else:
                print(f"Error: {response.status_code}, {response.text}")
        else:
            print(f'File {local_filename} already exists. If you want to download again delete the file')

        df = pd.read_excel(local_filename)
        dm = extract_hydro_capacity_at_year(df, yr)

        if dm_all is None:
            dm_all = dm.copy()
        else:
            dm_all.append(dm, dim='Years')

    dm_all.sort(dim='Years')

    dm_CH = dm_all.groupby({'Switzerland': '.*'}, regex=True, inplace=False, dim='Country')
    dm_all.append(dm_CH, dim='Country')

    return dm_all


def extract_old_hydro_capacity_zip(url_dict):

    dm_all = None
    for yr in url_dict.keys():

        df = pd.read_excel(url_dict[yr]['local_filename'])
        dm = extract_hydro_capacity_at_year(df, yr)

        if dm_all is None:
            dm_all = dm.copy()
        else:
            dm_all.append(dm, dim='Years')

    dm_all.sort(dim='Years')

    dm_CH = dm_all.groupby({'Switzerland': '.*'}, regex=True, inplace=False, dim='Country')
    dm_all.append(dm_CH, dim='Country')

    return dm_all


def extract_nuclear_capacity_data(reactor_list, dm_capacity):
    # 'Muhleberg': {'Canton': 'BE', 'Pmax': 373, 'StartYr': 1971, 'EndYr': 2019},

    dm = dm_capacity.filter({'Variables': ['pow_capacity-Pmax'], 'Categories1': ['Nuclear']})
    # Reset values to 0
    idx = dm.idx
    dm.array[...] = 0
    years_missing = list(set(range(dm.col_labels['Years'][0], dm.col_labels['Years'][-1])) - set(dm.col_labels['Years']))
    dm.add(0, dummy=True, dim='Years', col_label=years_missing)
    dm.sort('Years')
    for name, properties in reactor_list.items():
        cntr = properties['Canton']
        startyr = max(dm.col_labels['Years'][0], properties['StartYr'])
        endyr = min(dm.col_labels['Years'][-1], properties['EndYr'])
        for yr in range(startyr, endyr+1):
            dm.array[idx[cntr], idx[yr], idx['pow_capacity-Pmax'], idx['Nuclear']] += properties['Pmax']
    dm.array[idx['Switzerland'], ...] = np.nansum(dm.array, axis=0)
    return dm


def fill_missing_years_capacity_hydro(dm_capacity_hydro_ots, years_map, years_ots):
    years_missing = list(set(years_ots) - set(dm_capacity_hydro_ots.col_labels['Years']))
    dm_capacity_hydro_ots.add(np.nan, dim='Years', col_label=years_missing, dummy=True)
    idx = dm_capacity_hydro_ots.idx
    for ref_yr, yr_range in years_map.items():
        for yr in yr_range:
            dm_capacity_hydro_ots.array[:, idx[yr], ...] = dm_capacity_hydro_ots.array[:, idx[ref_yr], ...]

    dm_capacity_hydro_ots.sort('Years')
    return dm_capacity_hydro_ots


def adjust_based_on_nexuse(dm, dm_nexuse):
    dm.sort('Categories1')
    dm_nexuse.sort('Categories1')
    # Filter from Nexus-e the technologies in dm
    dm_filter = dm_nexuse.filter({'Country': ['Switzerland'], 'Years': years_ots, 'Variables': ['pow_capacity-Pmax'],
                                             'Categories1': dm.col_labels['Categories1']})

    # Join Nexus-e and dm (for Switzerland)
    dm.rename_col('pow_capacity-Pmax', 'pow_capacity-Pmax-old', dim='Variables')
    dm_tmp = dm.filter({'Country': ['Switzerland']})
    dm_filter.append(dm_tmp, dim='Variables')

    # Determine the ratio between Nexus-e data (Pmax) and our extracted data (Pmax-old)
    dm_filter.operation('pow_capacity-Pmax', '/', 'pow_capacity-Pmax-old', out_col='factor', unit='%')
    dm_filter.array[dm_filter.array == 0] = np.nan
    idx = dm_filter.idx
    # Compute the average adjustment throughout the years
    avg_factor = np.nanmean(dm_filter.array[:, :, idx['factor'], :], axis=1, keepdims=True)
    # Multiply Pmax (old) by the avg adjustment to match Nexus-e
    idx = dm.idx
    dm.array[:, :, idx['pow_capacity-Pmax-old'], :] = dm.array[:, :, idx['pow_capacity-Pmax-old'], :] * avg_factor
    dm.rename_col('pow_capacity-Pmax-old', 'pow_capacity-Pmax', dim='Variables')
    return dm


def extract_waste_capacity_ots(local_filename, years_ots):
    df = pd.read_excel(local_filename)
    df['Kanton'] = df['Kanton'].str.replace('\xa0', '')
    # Rename cols
    dict_rename = {'Kanton': 'Country', 'Täglicher Abfalldurchsatz (Tonnen)': 'ref_capacity-Pmax[t]',
                   'Elektrische Leistung (MW)': 'pow_capacity-Pmax[MW]', 'Inbetriebnahme': 'Years'}
    df.rename(dict_rename, axis=1, inplace=True)
    # Keep only important cols
    df = df[list(dict_rename.values())]
    all_years = range(df['Years'].min(), df['Years'].max() + 1)
    all_countries = df['Country'].unique()
    # 2. Create a MultiIndex for all combinations of countries and years
    multi_index = pd.MultiIndex.from_product([all_countries, all_years], names=['Country', 'Years'])
    # 3. Reindex the DataFrame with this MultiIndex
    df = df.set_index(['Country', 'Years']).reindex(multi_index)
    # 4. Fill missing values with 0
    df = df.fillna(0).astype(int)
    # Reset index if needed
    df = df.reset_index()
    dm = DataMatrix.create_from_df(df, num_cat=0)
    years_missing = list(set(years_ots) - set(dm.col_labels['Years']))
    dm.add(0, dummy=True, dim='Years', col_label=years_missing)
    dm.sort('Years')
    dm.array = np.cumsum(dm.array, axis=1)
    dm.filter({'Years': years_ots}, inplace=True)
    dm_CH = dm.groupby({'Switzerland': '.*'}, regex=True, dim='Country', inplace=False)
    dm.append(dm_CH, dim='Country')
    dm.rename_col(['pow_capacity-Pmax', 'ref_capacity-Pmax'], ['pow_capacity-Pmax_Waste', 'ref_capacity-Pmax_Waste'], dim='Variables')
    dm.deepen()
    return dm


def extract_oil_gas_capacity_data(local_filename):
    df = pd.read_excel(local_filename)
    dict_name = {'Inst. Leistung (MW)': 'Pmax', 'Brennstoff': 'Type'}
    df.rename(dict_name, axis=1, inplace=True)
    df = df[['Canton', 'Pmax', 'Type', 'StartYr', 'EndYr']]
    #df = df.pivot(columns='Type', index=['Canton', 'StartYr', 'EndYr'], values='Pmax')
    #df.reset_index(inplace=True)
    df['EndYr'] = df['EndYr'].fillna(2050)
    col_labels_dict = {'Country': list(df['Canton'].unique()),
                       'Years': years_ots,
                       'Variables': ['pow_capacity-Pmax'],
                       'Categories1': ['Oil', 'Gas']}
    dm = DataMatrix(col_labels_dict, units={'pow_capacity-Pmax': 'MW'})
    dm.array = np.zeros(tuple([len(list) for list in col_labels_dict.values()]))
    idx = dm.idx
    for yr in years_ots:
        for row in range(len(df)):
            StartYr = df.iloc[row]['StartYr']
            EndYr = df.iloc[row]['EndYr']
            if StartYr <= yr and yr <= EndYr:
                cntr = df.iloc[row]['Canton']
                cat = df.iloc[row]['Type']
                dm.array[idx[cntr], idx[yr], idx['pow_capacity-Pmax'], idx[cat]] += df.iloc[row]['Pmax']
    dm_CH = dm.groupby({'Switzerland': '.*'}, dim='Country', regex=True, inplace=False)
    dm.append(dm_CH, dim='Country')
    return dm


def extract_energy_statistics_data(file_url, local_filename, sheet_name, parameters):

    mapping = parameters['mapping']   # dictionary,  to rename column headers
    var_name = parameters['var name']  # string, dm variable name
    headers_idx = parameters['headers indexes']  # tuple with index of rows to keep for header
    first_row = parameters['first row']  # integer with the first row to keep # regex expression with cols to drop
    unit = parameters['unit']   # Put None if unit is in table, else str
    col_to_drop = parameters['cols to drop']   # None if no need to drop, else string (for dm.drop)

    if not os.path.exists(local_filename):
        response = requests.get(file_url, stream=True)
        # Check if the request was successful
        if response.status_code == 200:
            with open(local_filename, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            print(f"File downloaded successfully as {local_filename}")
        else:
            print(f"Error: {response.status_code}, {response.text}")
    else:
        print(f'File {local_filename} already exists. If you want to download again delete the file')

    df = read_excel_with_merged_cells(local_filename, sheet_name)
    combined_headers = []
    if unit is None:
        for col1, col2, col3 in zip(df.iloc[headers_idx[0]], df.iloc[headers_idx[1]], df.iloc[headers_idx[2]]):
            combined_headers.append(str(col1) + '-' + str(col2) + '[' + str(col3) + ']')
    else:
        for col1, col2 in zip(df.iloc[headers_idx[0]], df.iloc[headers_idx[1]]):
            combined_headers.append(str(col1) + '-' + str(col2) + '[' + unit + ']')
    # Set the new header
    df.columns = combined_headers
    df = df[first_row:].copy()

    def is_valid_number(val):
        return isinstance(val, (int, float)) and not pd.isna(val)

    # Apply the function to filter out rows with no valid numeric values
    df = df[df.apply(lambda row: row.map(is_valid_number).any(), axis=1)]
    # Apply similarly for columns if needed
    df = df.loc[:, df.apply(lambda col: col.map(is_valid_number).any())]

    df.rename({df.columns[0]: 'Years'}, axis=1, inplace=True)
    df['Country'] = 'Switzerland'
    df.replace('-', 0, inplace=True)
    dm = DataMatrix.create_from_df(df, num_cat=0)
    if col_to_drop is not None:
        dm.drop(dim='Variables', col_label=col_to_drop)

    for key in list(mapping.keys()):
        mapping[var_name + '_' + key] = mapping.pop(key)

    dm_out = dm.groupby(mapping, regex=True, dim='Variables', inplace=False)
    dm_out.deepen()
    dm_out.filter({'Years': years_ots}, inplace=True)
    #dm_out.change_unit(var_name, 277.8, 'TJ', 'MWh')

    return dm_out


def extract_districtheating_demand(file_url, local_filename, sheet_name, mapping, var_name):

    if not os.path.exists(local_filename):
        response = requests.get(file_url, stream=True)
        # Check if the request was successful
        if response.status_code == 200:
            with open(local_filename, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            print(f"File downloaded successfully as {local_filename}")
        else:
            print(f"Error: {response.status_code}, {response.text}")
    else:
        print(f'File {local_filename} already exists. If you want to download again delete the file')

    df = read_excel_with_merged_cells(local_filename, sheet_name)
    combined_headers = []
    for col1, col2 in zip(df.iloc[5], df.iloc[6]):
        combined_headers.append(str(col1) + '_' + str(col2) + '[TJ]')

    # Set the new header
    df.columns = combined_headers
    df = df[19:].copy()

    def is_valid_number(val):
        return isinstance(val, (int, float)) and not pd.isna(val)

    # Apply the function to filter out rows with no valid numeric values
    df = df[df.apply(lambda row: row.map(is_valid_number).any(), axis=1)]
    # Apply similarly for columns if needed
    df = df.loc[:, df.apply(lambda col: col.map(is_valid_number).any())]

    df.rename({'Année_Année[TJ]': 'Years'}, axis=1, inplace=True)
    df['Country'] = 'Switzerland'
    df.replace('-', 0, inplace=True)
    dm = DataMatrix.create_from_df(df, num_cat=0)
    dm.filter_w_regex({'Variables': 'Energie utilisé.*'}, inplace=True)
    dm.rename_col_regex('Energie utilisé_', '', dim='Variables')

    for key in list(mapping.keys()):
        mapping[var_name + '_' + key] = mapping.pop(key)

    dm_out = dm.groupby(mapping, regex=True, dim='Variables', inplace=False)
    dm_out.deepen()
    dm_out.filter({'Years': years_ots}, inplace=True)
    #dm_out.change_unit(var_name, 277.8, 'TJ', 'MWh')

    return dm_out


#years_setting = [1990, 2022, 2050, 5]  # Set the timestep for historical years & scenarios
years_ots = create_years_list(1990, 2023, 1)
years_fts = create_years_list(2025, 2050, 5)
baseyear = 2023

energy_url = 'https://www.uvek-gis.admin.ch/BFE/ogd/115/ogd115_gest_bilanz.csv'
local_path = 'data/Energy_statistic.csv'
outfile_dm = 'data/energy.pickle'

# Extract energy data
dm_energy = extract_energy_data(energy_url, local_path, baseyear, years_ots, outfile_dm)

# change units from TJ to TWh
for var in dm_energy.col_labels['Variables']:
    dm_energy.change_unit(var, factor=3600, old_unit='TJ', new_unit='TWh', operator='/')

# Some visualisation
plotting = False
if plotting:
    dm_energy.datamatrix_plot(title='All energy variables in TWh')
    dm_energy.datamatrix_plot({'Variables': 'Import'}, title='Import [TWh]')
    dm_energy.datamatrix_plot({'Variables': 'Energy conversion - Nuclear power plants'},
                              title='Energy conversion - Nuclear power plants [TWh]')
    dm_energy.datamatrix_plot({'Variables': 'Energy conversion - conventional-thermal power, district heating and district heating power plants'},
                              title='Energy conversion - conventional-thermal power, district heating and district heating power plants [TWh]')

# SECTION - Extract Capacity ots, from Nexus-e installed capacity (2020-2050)
file = 'data/Capacity_Nexuse.xlsx'
dm_capacity, dm_const = extract_nexuse_capacity_data(file)
dm_capacity_group = dm_capacity.copy()
dm_capacity_group.groupby({'Gas': 'Gas.*', 'Hydro': 'Dam|RoR'}, regex=True, dim='Categories1', inplace=True)

# SECTION - Extract Hydro Capacity ots, 1990-2020 from OFS
# Hydro-power
# Source:  https://www.bfe.admin.ch/bfe/fr/home/approvisionnement/energies-renouvelables/force-hydraulique.html#kw-96906
# 1991 - Statistik der Wasserkraftanlagen der Schweiz. Stand 1.1.1991
# 1996 - Statistik der Wasserkraftanlagen der Schweiz. Stand 1.1.1996
# 2001 - Statistik der Wasserkraftanlagen der Schweiz. Stand 1.1.2001
url_dict = {
            1990: {'file_url': 'https://www.bfe.admin.ch/bfe/fr/home/versorgung/erneuerbare-energien/wasserkraft.exturl.html/aHR0cHM6Ly9wdWJkYi5iZmUuYWRtaW4uY2gvZGUvcHVibGljYX/Rpb24vZG93bmxvYWQvNzkyNA==.html',
                   'local_filename': 'data/hydro_power_1990.xlsx'},
            1995: {'file_url': 'https://www.bfe.admin.ch/bfe/fr/home/versorgung/erneuerbare-energien/wasserkraft.exturl.html/aHR0cHM6Ly9wdWJkYi5iZmUuYWRtaW4uY2gvZGUvcHVibGljYX/Rpb24vZG93bmxvYWQvNzkyMw==.html',
                   'local_filename': 'data/hydro_power_1995.xlsx'},
            2000: {'file_url': 'https://www.bfe.admin.ch/bfe/fr/home/versorgung/erneuerbare-energien/wasserkraft.exturl.html/aHR0cHM6Ly9wdWJkYi5iZmUuYWRtaW4uY2gvZGUvcHVibGljYX/Rpb24vZG93bmxvYWQvNzkyMg==.html',
                   'local_filename': 'data/hydro_power_2000.xlsx'},
            2005: {'file_url': 'nan',
                   'local_filename': 'data/hydro_power_2005.xlsx'},
            2015: {'file_url': 'nan',
                   'local_filename': 'data/hydro_power_2015.xlsx'},
            2019: {'file_url': 'nan',
                   'local_filename': 'data/hydro_power_2019.xlsx'},
            2020: {'file_url': 'nan',
                   'local_filename': 'data/hydro_power_2020.xlsx'},
            2021: {'file_url': 'nan',
                   'local_filename': 'data/hydro_power_2021.xlsx'},
            2022: {'file_url': 'nan',
                   'local_filename': 'data/hydro_power_2022.xlsx'},
            2023: {'file_url': 'nan',
                   'local_filename': 'data/hydro_power_2023.xlsx'}
            }
dm_capacity_hydro_ots = extract_old_hydro_capacity_data(url_dict)
# Create step function profile
# Allocate missing years to existing year
years_map = {1990: range(1990, 1993+1), 1995: range(1994, 1997+1), 2000: range(1998, 2000+1),
             2005: range(2001, 2009+1), 2015: range(2010, 2017+1), 2019: range(2018, 2019+1)}
dm_capacity_hydro_ots = fill_missing_years_capacity_hydro(dm_capacity_hydro_ots, years_map, years_ots)

dm_capacity_hydro_ots = adjust_based_on_nexuse(dm_capacity_hydro_ots, dm_capacity)

# SECTION - Extract Nuclear Capacity ots, 1990-2020 from OFS
# Source:  https://de.wikipedia.org/wiki/Liste_der_Kernreaktoren_in_der_Schweiz
reactor_list = {'Muhleberg': {'Canton': 'BE', 'Pmax': 373, 'StartYr': 1971, 'EndYr': 2019},
                'Beznau1': {'Canton': 'AG', 'Pmax': 365, 'StartYr': 1969, 'EndYr': 2033},
                'Beznau2': {'Canton': 'AG', 'Pmax': 365, 'StartYr': 1971, 'EndYr': 2032},
                'Gosgen': {'Canton': 'SO', 'Pmax': 1010, 'StartYr': 1979, 'EndYr': 2060},
                'Liebstadt': {'Canton': 'AG', 'Pmax': 1233, 'StartYr': 1984, 'EndYr': 2060}}

dm_capacity_nuclear = extract_nuclear_capacity_data(reactor_list, dm_capacity)

# SECTION - Waste Capacity ots (power + heat)
# https://de.wikipedia.org/wiki/Liste_von_Kehrichtverbrennungsanlagen_in_der_Schweiz
local_filename = 'data/waste_power.xlsx'
# This has both the capacity in MW and the tonnes of waste incinerated every day (capacity)
dm_capacity_waste_ots = extract_waste_capacity_ots(local_filename, years_ots)

dm_capacity_waste_ots = adjust_based_on_nexuse(dm=dm_capacity_waste_ots, dm_nexuse=dm_capacity)

# SECTION - Oil & Gas Capacity ots (power)
local_filename = 'data/oil_gas_power_plants.xlsx'
dm_capacity_oilgas_ots = extract_oil_gas_capacity_data(local_filename)

# SECTION - Wind and Solar Capacity ots
# https://www.bfe.admin.ch/bfe/en/home/supply/statistics-and-geodata/energy-statistics/sector-statistics.html
# Excel file under Renewable Energy titled "Swiss Statistics of the Renewable energies" or
# "Schweizerische Statistik der erneuerbaren Energien 2023 - Datentabellen"
file_url = 'https://www.bfe.admin.ch/bfe/en/home/versorgung/statistik-und-geodaten/energiestatistiken/teilstatistiken.exturl.html/aHR0cHM6Ly9wdWJkYi5iZmUuYWRtaW4uY2gvZGUvcHVibGljYX/Rpb24vZG93bmxvYWQvODc4Nw==.html'
local_filename = 'data/swiss_statistics_of_the_renewable_energies.xlsx'
dm_capacity_PV_wind_ots = extract_renewable_capacity_data(file_url, local_filename)

# SECTION - Electricity Production
file_url = 'https://www.bfe.admin.ch/bfe/fr/home/versorgung/statistik-und-geodaten/energiestatistiken/gesamtenergiestatistik.exturl.html/aHR0cHM6Ly9wdWJkYi5iZmUuYWRtaW4uY2gvZnIvcHVibGljYX/Rpb24vZG93bmxvYWQvNzUxOQ==.html'
local_filename = 'data/statistique_globale_suisse_energie.xlsx'

dm_production = extract_production_data(file_url, local_filename)
dm_production.change_unit('pow_production', old_unit='MWh', new_unit='TWh', factor=1e-6)

# SECTION - Energy trade (incl.electricity)
# Extract import
mapping = {'wood': ".*Bois.*", 'biofuels': '.*biogènes.*', 'coal': '.*Charbon.*',
           'electricity': '.*Electricité.*', 'waste': '.*Ordures.*', 'gas': '.*Gaz.*', 'oil': '.*Pétrole.*'}
dm_import = extract_importexport_data(file_url, local_filename, sheet_name='T06', var_name='pow_import', mapping=mapping)
dm_import.change_unit('pow_import', old_unit='MWh', new_unit='TWh', factor=1e-6)
# Extract Export
mapping = {'wood': ".*Bois.*",  'coal': '.*Charbon.*',
           'electricity': '.*Electricité.*', 'oil': '.*pétroliers.*'}
dm_export = extract_importexport_data(file_url, local_filename, sheet_name='T07', var_name='pow_export', mapping=mapping)
dm_export.change_unit('pow_export', old_unit='MWh', new_unit='TWh', factor=1e-6)
# Compute Net Import
cat_missing = list(set(dm_import.col_labels['Categories1']) - set(dm_export.col_labels['Categories1']))
dm_export.add(0, col_label=cat_missing, dim='Categories1', dummy=True)
dm_import.append(dm_export, dim='Variables')
dm_import.operation('pow_import', '-', 'pow_export', out_col='pow_net-import', unit='TWh')

# Add net-import of electricity as electricity production import
dm_elec_import = dm_import.filter({'Variables': ['pow_net-import'], 'Categories1': ['electricity']})
dm_elec_import.rename_col('pow_net-import', 'pow_production', dim='Variables')
dm_elec_import.rename_col('electricity', 'Net-import', dim='Categories1')
dm_production.append(dm_elec_import, dim='Categories1')

# SECTION - Group Capacities
dm_CH = dm_capacity_oilgas_ots.filter({'Country': ['Switzerland']})
dm_CH.append(dm_capacity_waste_ots.filter({'Country': ['Switzerland'], 'Variables': ['pow_capacity-Pmax']}), dim='Categories1')
dm_CH.append(dm_capacity_nuclear.filter({'Country': ['Switzerland'], 'Years': years_ots}), dim='Categories1')
dm_CH.append(dm_capacity_hydro_ots.filter({'Country': ['Switzerland'], 'Years': years_ots}), dim='Categories1')
dm_CH.append(dm_capacity_PV_wind_ots.filter({'Country': ['Switzerland'], 'Years': years_ots}), dim='Categories1')


# SECTION - Split Gas into Gas-CC and Gas-GS based on Nexus-e
# Gas is originally Symple Cycle
dm_CH.rename_col('Gas', 'GasSC', dim='Categories1')
# But in recent years it becomes combined Cycle
for yr in dm_CH.col_labels['Years']:
    dm_CH['Switzerland', yr, 'pow_capacity-Pmax', 'GasSC'] = dm_CH['Switzerland', yr, 'pow_capacity-Pmax', 'GasSC'] \
                                                             - dm_capacity['Switzerland', yr, 'pow_capacity-Pmax', 'GasCC']
dm_tmp = dm_capacity.filter({'Country': ['Switzerland'], 'Years': years_ots, 'Variables': ['pow_capacity-Pmax'],
                             'Categories1': ['GasCC']})
dm_CH.append(dm_tmp, dim='Categories1')



# SECTION - Merge OTS capacity and FTS capacity in CH
missing_ots_cat = list(set(dm_capacity.col_labels['Categories1']) - set(dm_CH.col_labels['Categories1']))

dm_CH.add(0, dummy=True, dim='Categories1', col_label=missing_ots_cat)
dm_CH.sort('Categories1')
dm_capacity.sort('Categories1')
for yr in dm_CH.col_labels['Years']:
    dm_capacity['Switzerland', yr, 'pow_capacity-Pmax', :] = dm_CH['Switzerland', yr, 'pow_capacity-Pmax', :]
# Avoid 2025 discontinuity (especially in PV)
dm_capacity['Switzerland', 2025, 'pow_capacity-Pmax', :] = np.nan
dm_capacity.fill_nans('Years')

# SECTION - Add installed capacity to pickle
dm_capacity.add(np.nan, dim='Variables', col_label='pow_existing-capacity', dummy=True, unit='MW')
for yr in years_ots:
    dm_capacity['Switzerland', yr, 'pow_existing-capacity', :] = dm_capacity['Switzerland', yr, 'pow_capacity-Pmax', :]
dm_capacity.fill_nans('Years')
for yr in years_fts:
    dm_capacity['Switzerland', yr, 'pow_existing-capacity', :] = np.minimum(dm_capacity['Switzerland', yr, 'pow_existing-capacity', :],
                                                                            dm_capacity['Switzerland', yr, 'pow_capacity-Pmax', :])


# SECTION - Determine split Oil-Gas into GasCC, GasSC, Oil and Cogen using capacity factors
dm_prod_oilgas = dm_production.filter({'Country': ['Switzerland'], 'Categories1': ['Oil-Gas'],
                                       'Variables': ['pow_production']})
dm_cap_oilgas = dm_capacity.filter_w_regex({'Variables': 'pow_existing-capacity', 'Categories1': ".*Gas.*|Oil",
                                            'Country': 'Switzerland'})
dm_cap_oilgas.filter({'Years': years_ots}, inplace=True)
dm_cap_oilgas.normalise('Categories1', inplace=True, keep_original=False)
arr_prod_split = dm_cap_oilgas.array * dm_prod_oilgas.array
dm_production.add(arr_prod_split, dim='Categories1', col_label=dm_cap_oilgas.col_labels['Categories1'])
dm_production.drop('Categories1', 'Oil-Gas')
cat_matching = list(set(dm_capacity.col_labels['Categories1']).intersection(set(dm_production.col_labels['Categories1'])))
dm_cap_match = dm_capacity.filter({'Categories1': cat_matching, 'Country': ['Switzerland'],
                                   'Variables': ['pow_existing-capacity'], 'Years': years_ots})
dm_prod_match = dm_production.filter({'Categories1': cat_matching})
dm_cap_match.append(dm_prod_match, dim='Variables')
dm_cap_match.change_unit('pow_existing-capacity', old_unit='MW', new_unit='GW', factor=1e-3)
dm_cap_match.operation('pow_production', '/', 'pow_existing-capacity', out_col='pow_cap-fact', unit='TWh/GW')
dm_cap_match.change_unit('pow_cap-fact', old_unit='TWh/GW', new_unit='%', factor=8.760, operator='/')


mask = dm_cap_match[:, :, 'pow_cap-fact', :] > 1
dm_cap_match[:, :, 'pow_cap-fact', :][mask] = np.nan
dm_cap_match.fill_nans('Years')

dm_cap_match.change_unit('pow_cap-fact', old_unit='%', new_unit='TWh/GW', factor=8.760, operator='*')
dm_cap_match.operation('pow_production', '/', 'pow_cap-fact', out_col='pow_capacity', unit='GW')
dm_cap_match.change_unit('pow_capacity', old_unit='GW', new_unit='MW', factor=1000)

dm_cap_match.filter({'Variables': ['pow_capacity']}, inplace=True)

dm_cap_ots = dm_capacity.filter({'Years': years_ots})
dm_capacity.drop(dim='Years', col_label=years_ots)
for cat in dm_cap_match.col_labels['Categories1']:
    dm_cap_ots['Switzerland', :, 'pow_existing-capacity', cat] = dm_cap_match['Switzerland', :, 'pow_capacity', cat]
dm_capacity.append(dm_cap_ots, dim='Years')
dm_capacity.sort('Years')

# !FIXME: The ots capacity data for example for oil and gas are available by canton,
# There is maybe a problem with the Nexus-e capacities
# SECTION - Redistribute Swiss capacity-Pmax and existing-capacity by Canton
dm_capacity.drop(dim='Variables', col_label=['pow_capacity-Emax', 'pow_capacity-Pmin'])
dm_capacity_CH = dm_capacity.filter({'Country': ['Switzerland']})
dm_capacity.drop(dim='Country', col_label='Switzerland')
dm_capacity[:, :, 'pow_existing-capacity', :] = dm_capacity[:, :, 'pow_capacity-Pmax', :]
dm_capacity.fill_nans('Years')
dm_capacity.normalise(inplace=True, dim='Country')
dm_capacity.array = dm_capacity[:, :, :, :]*dm_capacity_CH['Switzerland', np.newaxis, :, :, :]
for var in dm_capacity.col_labels['Variables']:
    dm_capacity.units[var] = dm_capacity_CH.units[var]

mask = np.isnan(dm_capacity[:, :, :, :])
dm_capacity[:, :, :, :][mask] = 0

dm_capacity.append(dm_capacity_CH, dim='Country')


######################################################################
##  Extract historical demand of fuels other than electricity prod  ##
######################################################################

# Statistique Global Suisse de l'Energie
file_url = 'https://www.bfe.admin.ch/bfe/fr/home/versorgung/statistik-und-geodaten/energiestatistiken/gesamtenergiestatistik.exturl.html/aHR0cHM6Ly9wdWJkYi5iZmUuYWRtaW4uY2gvZnIvcHVibGljYX/Rpb24vZG93bmxvYWQvNzUxOQ==.html'
local_filename = 'data/statistique_globale_suisse_energie.xlsx'
# Bois et charbon de bois1
# Force hydraulique
# Ordures ménagères et déchets industriels2
# Charbon
# Pétrole brut et produits pétroliers
# dont pétrole brut
# dont produits pétroliers
# Gaz
# Combustibles nucléaires
# Autres énergies renouvelables3
# Utilisation totale d'agents énergé-tiques
# Elektricité Solde de Import/Export
# Consommation brute d'énergie dans le pays (100%)
parameters = dict()
mapping = {'heating-oil': '.*Combustibles.*',  'transport-oil': '.*Carburants.*',  'gas': '.*Gaz.*',
           'coal': '.*Charbon.*', 'wood': '.*bois.*', 'district-heating': '.*distance.*', 'waste': '.*industriel.*',
           'biofuels': '.*biogènes.*', 'biogas': '.*Biogaz.*'}
parameters['mapping'] = mapping  # dictionary,  to rename column headers
parameters['var name'] = 'pow_fuel-demand'  # string, dm variable name
parameters['headers indexes'] = (5, 6)  # tuple with index of rows to keep for header
parameters['first row'] = 87  # integer with the first row to keep
parameters['unit'] = 'TJ'
parameters['cols to drop'] = None

dm_fuels_demand = extract_energy_statistics_data(file_url, local_filename, sheet_name='T14', parameters=parameters)


##### Supply
# SECTION - Energy supply
parameters = dict()
mapping = {'hydro-power': '.*hydraulique.*', 'wood': '.*bois.*', 'waste': '.*déchets.*', 'coal': '.*charbon.*',
           'oil': '.*Pétrole brut et produits pétroliers.*', 'gas': '.*Gaz.*', 'nuclear': '.*nucléaires.*',
           'renewables': '.*renouvelables.*'}
parameters['mapping'] = mapping  # dictionary,  to rename column headers
parameters['var name'] = 'pow_fuel-supply'  # string, dm variable name
parameters['headers indexes'] = (4, 5, 5)  # tuple with index of rows to keep for header
parameters['first row'] = 86  # integer with the first row to keep
parameters['cols to drop'] = '.*%.*'
parameters['unit'] = None

dm_fuels_supply = extract_energy_statistics_data(file_url, local_filename, sheet_name='T10', parameters=parameters)


##### Losses
# SECTION - Losses of electricity
parameters = dict()
mapping = {'Losses': '.*Centrales électriques.*'}
parameters['mapping'] = mapping  # dictionary,  to rename column headers
parameters['var name'] = 'pow_production'  # string, dm variable name
parameters['headers indexes'] = (3, 4)  # tuple with index of rows to keep for header
parameters['first row'] = 5  # integer with the first row to keep
parameters['cols to drop'] = '.*Raffineries.*|.*Usines.*|.*Chaleur.*|.*Total.*|.*Consommation.*|.*%.*'
parameters['unit'] = 'TJ'

dm_losses = extract_energy_statistics_data(file_url, local_filename, sheet_name='T13', parameters=parameters)
dm_losses.change_unit('pow_production', factor=3600, old_unit='TJ', new_unit='TWh', operator='/')
# Remove Pump-Open losses from total losses to avoid double counting
dm_losses['Switzerland', :, 'pow_production', 'Losses'] = dm_losses['Switzerland', :, 'pow_production', 'Losses'] - dm_production['Switzerland', :, 'pow_production', 'Pump-Open']
dm_losses.array = -dm_losses.array
dm_production.append(dm_losses, dim='Categories1')

#######################################
###      Oil Supply by type     #######
#######################################
mapping = {'heating-oil': '.*Huile.*', 'kerosene': '.*aviation.*', 'diesel': '.*diesel.*', 'gasoline': '.*Essence2-Total.*',
           'other': '.*Coke.*|.*Autres.*'}
parameters['mapping'] = mapping  # dictionary,  to rename column headers
parameters['var name'] = 'pow_fuel-supply'  # string, dm variable name
parameters['headers indexes'] = (5, 6)  # tuple with index of rows to keep for header
parameters['first row'] = 23  # integer with the first row to keep
parameters['cols to drop'] = None
parameters['unit'] = '1000t'
dm_oil_split = extract_energy_statistics_data(file_url, local_filename, sheet_name='T20', parameters=parameters)
# Conversion factor based on Lower Heating Value
# Reference https://world-nuclear.org/information-library/facts-and-figures/heat-values-of-various-fuels
# Reference heating-oil https://www.forestresearch.gov.uk/tools-and-resources/fthr/biomass-energy-resources/reference-biomass/facts-figures/typical-calorific-values-of-fuels/
# Reference kerosene Linstrom, Peter (2021). NIST Chemistry WebBook. NIST Standard Reference Database Number 69. NIST Office of Data and Informatics. doi:10.18434/T4D303.
# (value reported by wikipedia https://en.wikipedia.org/wiki/Heat_of_combustion#cite_note-NIST-11)
# For other I'm using 42.
LHV_MJ_kg = {'heating-oil': 42.5, 'diesel': 44, 'gasoline': 45, 'kerosene': 44.1, 'other':  42}
# MJ/kg = TJ/1000tonnes
for var, conv_fact in LHV_MJ_kg.items():
    dm_oil_split[:, :, :, var] = conv_fact * dm_oil_split[:, :, :, var]
dm_oil_split.change_unit('pow_fuel-supply', factor=1, old_unit='1000t', new_unit='TJ')


# Map oil supply using dm_oil_split
dm_fuels_supply.append(dm_oil_split, dim='Categories1')
dm_fuels_supply.drop(dim='Categories1', col_label=['oil'])
keep_fuel_cat = ['gasoline', 'gas', 'waste', 'heating-oil', 'wood', 'diesel', 'kerosene']
dm_fuels_supply.filter({'Categories1': keep_fuel_cat}, inplace=True)
dm_fuels_supply.change_unit('pow_fuel-supply', factor=3600, old_unit='TJ', new_unit='TWh', operator='/')

# District-heating
#mapping = {'wood': 'Bois', 'coal': 'Charbon', 'nuclear': 'Combustibles nucléaires3', 'other': 'Divers4',
#           'electricity': 'Electricité', 'gas': 'Gaz1', 'heating-oil': 'Huile extra-légère|Huile moyenne et lourde',
#           'waste': 'Ordures2'}
#dm_distrheat = extract_districtheating_demand(file_url, local_filename, sheet_name='T26', mapping=mapping,
#                                              var_name='pow_district-heating')


#  but they are not appearing correctly in the final dm_capacity
# Correct this. Then I think the rest is good.

file = '../../../data/datamatrix/energy.pickle'
with open(file, 'rb') as handle:
    DM_energy = pickle.load(handle)

dm_capacity.rename_col('VD', 'Vaud', dim='Country')
DM_energy['capacity'] = dm_capacity.filter({'Country': ['Switzerland', 'Vaud']})
DM_energy['production'] = dm_production.filter({'Country': ['Switzerland']})
DM_energy['fuels'] = dm_fuels_supply

file = '../../../data/datamatrix/energy.pickle'
#with open(file, 'wb') as handle:
#    pickle.dump(DM_energy, handle, protocol=pickle.HIGHEST_PROTOCOL)
my_pickle_dump(DM_energy, file)

print('Hello')
