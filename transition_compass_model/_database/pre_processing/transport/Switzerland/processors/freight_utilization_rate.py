"""
Swiss freight OTS: load factor and utilisation rate for road modes (HDVH, HDVL, HDVM).

Variables
---------
tra_freight_load-factor : tkm/vkm
    Average payload per km driven.
    HDVH (TT/AT): BFS GTS HAV total tkm / EP2050 TT/AT VKM share.
    HDVM (RigidTruck): BFS GTS LORRY total tkm / EP2050 RigidTruck VKM share.
    HDVL (LCV): fixed at _HDVL_PAYLOAD_T_PER_VKM (0.2 t/vkm) by construction —
        HDVL tkm = LCV VKM × 0.2 in freight_tkm_modal_share.

tra_freight_utilisation-rate : vkm/year
    km driven per vehicle per year.
    HDVH: EP2050 HGV VKM × fleet_fraction / TT/AT fleet count.
    HDVM: EP2050 HGV VKM × fleet_fraction / RigidTruck fleet count.
    HDVL: EP2050 LCV VKM / LCV fleet count.

HGV VKM is split between HDVH and HDVM in proportion to their fleet counts,
since EP2050 does not disaggregate HGV VKM by sub-type.
BFS GTS tkm uses total traffic (CH + foreign registrations) to capture all
freight demand in Switzerland, while EP2050 covers all road activity including transit.

Sources
-------
EP2050 Verkehrssektor ZERO-Basis:
  Sheet 03 Fahrleistung: LCV and HGV total VKM (Mio. Verkehrs-km), 1990-2060.
  Sheet 02 Flottenbestand: fleet counts by segment, 1990-2060.
  # https://www.bfe.admin.ch/bfe/en/home/policy/energy-perspectives-2050-plus.html
  # file: EP2050+_Detailergebnisse 2020-2060_Verkehrssektor_alle Szenarien_2022-04-12.xlsx

BFS GTS (HAV, LORRY) total tkm (all registrations):
  # https://www.bfs.admin.ch/bfs/fr/home/statistiques/mobilite-transports/
  #   transport-marchandises/route/vehicules-lourds.html
  # file: ts-x-11.05-GTS-E26.csv
"""

import os

import numpy as np
import openpyxl
import pandas as pd

from transition_compass_model.model.common.data_matrix_class import DataMatrix

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
# ASTRA Swiss road freight statistics: LCV average load ≈ 20 % of ~1 t capacity.
# https://www.astra.admin.ch/astra/de/home/themen/Schwerverkehrsabgabe/vollzug-svag-lsva/schwerverkehrsstatistik.html
_HDVL_LOAD_FACTOR = 0.2  # tkm/vkm, set by construction in freight_tkm_modal_share

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR = os.path.join(_THIS_DIR, "../data")
_EP2050_PATH = os.path.join(
    _DATA_DIR,
    "EP2050_sectors/EP2050+_Szenarienergebnisse_Details_Nachfragesektoren/"
    "EP2050+_Detailergebnisse 2020-2060_Verkehrssektor_alle Szenarien_2022-04-12.xlsx",
)

_MODES_ROAD = ["HDVH", "HDVL", "HDVM"]


# ---------------------------------------------------------------------------
# Readers
# ---------------------------------------------------------------------------
def _read_ep2050_vkm() -> dict:
    """Return {mode: pd.Series(year->million_vkm)} for LCV and HGV from EP2050 Fahrleistung."""
    wb = openpyxl.load_workbook(_EP2050_PATH, read_only=True, data_only=True)
    ws = wb["03 Fahrleistung"]
    rows = list(ws.iter_rows(min_row=20, max_row=300, values_only=True))
    header = rows[0]
    years_ep = [int(v) for v in header[4:] if isinstance(v, (int, float))]
    n = len(years_ep)

    lcv = np.zeros(n)
    hgv = np.zeros(n)
    in_lcv = False
    in_hgv = False
    done_lcv = False
    done_hgv = False

    for r in rows[1:]:
        if r[1] == "LCV" and not done_lcv:
            vals = [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n]]
            lcv += np.array(vals)
            in_lcv = True
        elif in_lcv and r[1] != "LCV":
            done_lcv = True

        if r[1] == "HGV" and not done_hgv:
            vals = [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n]]
            hgv += np.array(vals)
            in_hgv = True
        elif in_hgv and r[1] != "HGV":
            done_hgv = True

        if done_lcv and done_hgv:
            break

    return {
        "LCV": pd.Series(dict(zip(years_ep, lcv))),
        "HGV": pd.Series(dict(zip(years_ep, hgv))),
    }


def _read_ep2050_fleet() -> dict:
    """Return {mode: pd.Series(year->count)} for TT/AT, RigidTruck, LCV from Flottenbestand.

    Reads only the first (ZERO-Basis) sub-table. The sheet contains three scenario
    sub-tables (ZERO-Basis, ZERO-A, ZERO-B); reading stops at the first scenario label.
    """
    wb = openpyxl.load_workbook(_EP2050_PATH, read_only=True, data_only=True)
    ws = wb["02 Flottenbestand"]
    rows = list(ws.iter_rows(min_row=20, max_row=300, values_only=True))
    header = rows[0]
    years_ep = [int(v) for v in header[4:] if isinstance(v, (int, float))]
    n = len(years_ep)

    hdvh = np.zeros(n)
    hdvm = np.zeros(n)
    hdvl = np.zeros(n)

    for r in rows[1:]:
        # Stop at next scenario boundary (row label like "ZERO-A")
        if r[0] is not None and isinstance(r[0], str) and r[0].startswith("ZERO"):
            break
        if r[1] is None:
            continue
        vals = np.array(
            [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n]]
        )
        if r[1].startswith("TT/AT"):
            hdvh += vals
        elif r[1].startswith("RigidTruck"):
            hdvm += vals
        elif r[1].startswith("LCV"):
            hdvl += vals

    return {
        "HDVH": pd.Series(dict(zip(years_ep, hdvh))),
        "HDVM": pd.Series(dict(zip(years_ep, hdvm))),
        "HDVL": pd.Series(dict(zip(years_ep, hdvl))),
    }


def _read_bfs_gts_road() -> dict:
    """Return {mode: pd.Series(year->million_tkm)} for HDVH and HDVM from BFS GTS."""
    path = os.path.join(_DATA_DIR, "Freight/ts-x-11.05-GTS-E26.csv")
    df = pd.read_csv(path, sep=";")
    df = df[(df["IMMATRICULATION"] == "_T") & (df["TRAFFIC_TYPE"] == "_T")]
    df = df[df["VEH_TYPE"].isin(["LORRY", "HAV"])]
    piv = df.pivot(index="REF_YEAR", columns="VEH_TYPE", values="OBS_VALUE")
    piv.index = piv.index.astype(int)
    return {"HDVH": piv["HAV"], "HDVM": piv["LORRY"]}


def _fill(raw: pd.Series, years: list) -> np.ndarray:
    s = raw.reindex(years, fill_value=np.nan).astype(float)
    s = s.interpolate(method="index")
    return s.bfill().ffill().values


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def run(years_ots: list, country_list: list) -> DataMatrix:
    """Build freight_utilization-rate DataMatrix for road modes (OTS only).

    Parameters
    ----------
    years_ots : list of int
    country_list : list of str

    Returns
    -------
    DataMatrix with Variables=['tra_freight_load-factor', 'tra_freight_utilisation-rate'],
    Categories1=['HDVH', 'HDVL', 'HDVM'], covering years_ots.
    """
    # --- Load raw data -------------------------------------------------------
    ep_vkm = _read_ep2050_vkm()  # million vkm per year
    ep_fleet = _read_ep2050_fleet()  # vehicle count per year
    bfs_tkm = _read_bfs_gts_road()  # million tkm per year

    # Fill all series to OTS years
    lcv_vkm = _fill(ep_vkm["LCV"], years_ots)  # million vkm
    hgv_vkm = _fill(ep_vkm["HGV"], years_ots)  # million vkm
    hdvh_fleet = _fill(ep_fleet["HDVH"], years_ots)
    hdvm_fleet = _fill(ep_fleet["HDVM"], years_ots)
    hdvl_fleet = _fill(ep_fleet["HDVL"], years_ots)
    hdvh_tkm = _fill(bfs_tkm["HDVH"], years_ots)  # million tkm
    hdvm_tkm = _fill(bfs_tkm["HDVM"], years_ots)  # million tkm

    # --- Split HGV VKM between HDVH and HDVM by fleet fraction ---------------
    total_hgv_fleet = hdvh_fleet + hdvm_fleet
    frac_hdvh = np.where(total_hgv_fleet > 0, hdvh_fleet / total_hgv_fleet, 0.5)
    hdvh_vkm = hgv_vkm * frac_hdvh  # million vkm
    hdvm_vkm = hgv_vkm * (1 - frac_hdvh)  # million vkm

    # --- Load factor (tkm/vkm) ------------------------------------------------
    # tkm and vkm are both in millions so the ratio is dimensionless (tkm/vkm)
    lf_hdvh = np.where(hdvh_vkm > 0, hdvh_tkm / hdvh_vkm, np.nan)
    lf_hdvm = np.where(hdvm_vkm > 0, hdvm_tkm / hdvm_vkm, np.nan)
    lf_hdvl = np.full(len(years_ots), _HDVL_LOAD_FACTOR)

    # --- Utilisation rate (vkm/vehicle/year) ----------------------------------
    # Convert million vkm → absolute vkm, divide by fleet count
    ur_hdvh = np.where(hdvh_fleet > 0, hdvh_vkm * 1e6 / hdvh_fleet, np.nan)
    ur_hdvm = np.where(hdvm_fleet > 0, hdvm_vkm * 1e6 / hdvm_fleet, np.nan)
    ur_hdvl = np.where(hdvl_fleet > 0, lcv_vkm * 1e6 / hdvl_fleet, np.nan)

    # --- Build DataMatrix -----------------------------------------------------
    all_countries = ["Switzerland", "Vaud"] + [
        c for c in country_list if c not in ("Switzerland", "Vaud")
    ]

    dm = DataMatrix(
        col_labels={
            "Country": all_countries,
            "Years": years_ots,
            "Variables": ["tra_freight_load-factor", "tra_freight_utilisation-rate"],
            "Categories1": _MODES_ROAD,
        },
        units={
            "tra_freight_load-factor": "tkm/vkm",
            "tra_freight_utilisation-rate": "vkm/year",
        },
    )
    idx = dm.idx
    ch = "Switzerland"

    data_lf = {"HDVH": lf_hdvh, "HDVL": lf_hdvl, "HDVM": lf_hdvm}
    data_ur = {"HDVH": ur_hdvh, "HDVL": ur_hdvl, "HDVM": ur_hdvm}

    for mode in _MODES_ROAD:
        dm.array[idx[ch], :, idx["tra_freight_load-factor"], idx[mode]] = data_lf[mode]
        dm.array[idx[ch], :, idx["tra_freight_utilisation-rate"], idx[mode]] = data_ur[
            mode
        ]

    # utilisation-rate is vehicle-type specific, not canton-specific → copy CH to all
    for country in all_countries:
        if country != ch:
            dm.array[idx[country], :, :, :] = dm.array[idx[ch], :, :, :]

    return dm
