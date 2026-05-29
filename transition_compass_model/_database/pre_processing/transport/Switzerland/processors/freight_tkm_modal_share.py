"""
Swiss freight OTS: total tkm and modal share by mode (1990-2023).

Sources
-------
Road HDVH (HAV) + HDVM (LORRY):
  BFS GTS survey (annual million tkm, 1993-2024, all registrations + all traffic types).
  LORRY = rigid trucks ≥ 3.5 t → HDVM; HAV = articulated trucks → HDVH.
  # https://www.bfs.admin.ch/bfs/fr/home/statistiques/mobilite-transports/
  #   transport-marchandises/route/vehicules-lourds.html
  # file: ts-x-11.05-GTS-E26.csv

Road HDVL (light commercial vehicles < 3.5 t):
  EP2050+ ZERO-Basis scenario LCV VKM (Mio VKM, 1990-2060), converted with
  _HDVL_PAYLOAD_T_PER_VKM (ASTRA average load factor ~20 % of ~1 t capacity).
  # https://www.bfe.admin.ch/bfe/en/home/policy/energy-perspectives-2050-plus.html
  # file: EP2050+_Detailergebnisse 2020-2060_Verkehrssektor_alle Szenarien_2022-04-12.xlsx
  #   sheet '03 Fahrleistung', ZERO-Basis scenario

Rail freight:
  BFS public transport statistics, table T7.2.1 row 12 (Chemins de fer, million tkm).
  Years 1990, 1995, 2000-2023; gaps 1991-1994 and 1996-1999 interpolated linearly.
  # https://www.bfs.admin.ch/bfs/fr/home/statistiques/mobilite-transports/
  #   transports-publics.html
  # file: tra_public_transport.xlsx, sheet T7_2_1

IWW (Rhine navigation):
  BFS Rhine imports/exports (million tonnes, 1955-2024) × _IWW_HAUL_KM.
  _IWW_HAUL_KM = 34 km is the navigable Swiss Rhine stretch (Port of Basel to
  German border), which reproduces the ~0.17 bn-tkm IWW share in Swiss statistics.
  # https://www.bfs.admin.ch/bfs/en/home/statistics/mobility-transport/
  #   goods-transport/air-water-pipeline.html
  # file: gr-e-11.05.04b.csv

Marine:
  Hardcoded 0 (Switzerland landlocked).

Aviation:
  Residual after all other modes to maintain _AVIATION_SHARE = 4.4 % of total.
  Calibrated to BAZL Swiss airport cargo (~500 kt/yr at ~2600 km avg haul ≈ 1.3 bn-tkm).

Vaud canton share:
  BFS GTS canton survey 2016-2020 average. Freight originating from Vaud: 767 635
  out of 11 013 466 thousand tkm = 6.97 %. Applied to all modes; IWW = 0 for Vaud
  (Vaud canton has no access to the navigable Rhine).
  # https://www.bfs.admin.ch/bfs/fr/home/statistiques/mobilite-transports/transport-marchandises/route/vehicules-lourds.assetdetail.18924747.html
  # file: su-f-11.05-GTS16-20-P19.xlsx, sheet FR
"""

import os

import numpy as np
import openpyxl
import pandas as pd

from transition_compass_model.model.common.data_matrix_class import DataMatrix

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
_HDVL_PAYLOAD_T_PER_VKM = (
    0.2  # ASTRA LCV avg load ≈ 20 % of 1 t capacity; see module docstring
)
_IWW_HAUL_KM = (
    34.0  # navigable Swiss Rhine (Port of Basel to German border); see module docstring
)
_AVIATION_SHARE = (
    0.044  # residual share calibrated to BAZL airport cargo data; see module docstring
)
_VAUD_ROAD_FREIGHT_SHARE = (
    767_635 / 11_013_466
)  # BFS GTS canton survey 2016-2020; see module docstring

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR = os.path.join(_THIS_DIR, "../data")

_MODES = ["HDVH", "HDVL", "HDVM", "IWW", "aviation", "marine", "rail"]


# ---------------------------------------------------------------------------
# Private readers
# ---------------------------------------------------------------------------
def _read_bfs_gts_road():
    """Return dict {mode: pd.Series(year→million tkm)} for HDVH and HDVM."""
    path = os.path.join(_DATA_DIR, "Freight/ts-x-11.05-GTS-E26.csv")
    df = pd.read_csv(path, sep=";")
    df = df[(df["IMMATRICULATION"] == "_T") & (df["TRAFFIC_TYPE"] == "_T")]
    df = df[df["VEH_TYPE"].isin(["LORRY", "HAV"])]
    piv = df.pivot(index="REF_YEAR", columns="VEH_TYPE", values="OBS_VALUE")
    piv.index = piv.index.astype(int)
    return {
        "HDVH": piv["HAV"],
        "HDVM": piv["LORRY"],
    }


def _read_ep2050_lcv_vkm():
    """Return pd.Series(year→million tkm) for HDVL using EP2050 LCV VKM × payload."""
    path = os.path.join(
        _DATA_DIR,
        "EP2050_sectors/EP2050+_Szenarienergebnisse_Details_Nachfragesektoren/"
        "EP2050+_Detailergebnisse 2020-2060_Verkehrssektor_alle Szenarien_2022-04-12.xlsx",
    )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["03 Fahrleistung"]
    rows = list(ws.iter_rows(min_row=20, max_row=350, values_only=True))
    header = rows[0]
    years_ep = [int(v) for v in header[4:] if isinstance(v, (int, float))]

    lcv_vkm = np.zeros(len(years_ep))
    in_lcv = False
    done = False
    for r in rows[1:]:
        if r[1] == "LCV" and not done:
            lcv_vkm += np.array(
                [
                    v if isinstance(v, (int, float)) else 0.0
                    for v in r[4 : 4 + len(years_ep)]
                ]
            )
            in_lcv = True
        elif in_lcv and r[1] != "LCV":
            done = True

    series = pd.Series(
        {yr: lcv_vkm[i] * _HDVL_PAYLOAD_T_PER_VKM for i, yr in enumerate(years_ep)}
    )
    return series


def _read_rail_tkm():
    """Return pd.Series(year→million tkm) for rail from BFS T7.2.1."""
    path = os.path.join(_DATA_DIR, "tra_public_transport.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["T7_2_1"]
    row4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    year_cols = [(i, int(v)) for i, v in enumerate(row4) if isinstance(v, (int, float))]
    row12 = list(ws.iter_rows(min_row=12, max_row=12, values_only=True))[0]

    data = {yr: row12[i] for i, yr in year_cols if isinstance(row12[i], (int, float))}
    return pd.Series(data, dtype=float)


def _read_iww_tkm():
    """Return pd.Series(year→million tkm) for IWW (Rhine tonnes × haul factor)."""
    path = os.path.join(_DATA_DIR, "Freight/gr-e-11.05.04b.csv")
    df = pd.read_csv(path, names=["Year", "total_Mt", "rhine_Mt"], skiprows=1)
    df["Year"] = df["Year"].astype(int)
    df = df.set_index("Year")
    return df["rhine_Mt"] * _IWW_HAUL_KM


def _fill_series(raw: pd.Series, years_ots: list) -> np.ndarray:
    """Reindex to years_ots, interpolate interior gaps, bfill/ffill ends."""
    s = raw.reindex(years_ots, fill_value=np.nan).astype(float)
    s = s.interpolate(method="index")
    s = s.bfill().ffill()
    return s.values


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def run(years_ots: list, country_list: list) -> dict:
    """Build freight_tkm and freight_modal-share DataMatrix for OTS.

    Parameters
    ----------
    years_ots : list of int
        Annual years from start (e.g. 1990) to base year (e.g. 2023).
    country_list : list of str
        Additional countries beyond Switzerland and Vaud to populate
        by duplicating Switzerland values (e.g. ['Vaud']).

    Returns
    -------
    dict with keys 'freight_tkm' and 'freight_modal-share', each a DataMatrix.
    """
    n_yr = len(years_ots)

    # --- Load raw series --------------------------------------------------
    gts = _read_bfs_gts_road()
    hdvh_raw = gts["HDVH"]
    hdvm_raw = gts["HDVM"]
    hdvl_raw = _read_ep2050_lcv_vkm()
    rail_raw = _read_rail_tkm()
    iww_raw = _read_iww_tkm()

    # --- Fill to OTS years ------------------------------------------------
    ch = {
        "HDVH": _fill_series(hdvh_raw, years_ots),
        "HDVM": _fill_series(hdvm_raw, years_ots),
        "HDVL": _fill_series(hdvl_raw, years_ots),
        "rail": _fill_series(rail_raw, years_ots),
        "IWW": _fill_series(iww_raw, years_ots),
        "marine": np.zeros(n_yr),
    }

    # Aviation = residual to hold _AVIATION_SHARE of total
    subtotal = sum(ch[m] for m in ch)
    ch["aviation"] = subtotal * (_AVIATION_SHARE / (1.0 - _AVIATION_SHARE))
    total_ch = subtotal + ch["aviation"]

    # --- Vaud: scale by canton share, IWW = 0 ----------------------------
    vd = {}
    for m in _MODES:
        if m in ("IWW", "marine"):
            vd[m] = np.zeros(n_yr)
        else:
            vd[m] = ch[m] * _VAUD_ROAD_FREIGHT_SHARE
    # Recompute Vaud aviation as residual with the same share
    vd_subtotal = sum(vd[m] for m in vd if m != "aviation")
    vd["aviation"] = vd_subtotal * (_AVIATION_SHARE / (1.0 - _AVIATION_SHARE))
    total_vd = vd_subtotal + vd["aviation"]

    # --- DataMatrix: freight_modal-share ---------------------------------
    all_countries = ["Switzerland", "Vaud"] + [
        c for c in country_list if c not in ("Switzerland", "Vaud")
    ]
    n_c = len(all_countries)
    n_m = len(_MODES)

    dm_modal = DataMatrix(
        col_labels={
            "Country": all_countries,
            "Years": years_ots,
            "Variables": ["tra_freight_modal-share"],
            "Categories1": _MODES,
        },
        units={"tra_freight_modal-share": "%"},
    )
    idx = dm_modal.idx

    for j, yr in enumerate(years_ots):
        for m in _MODES:
            dm_modal.array[idx["Switzerland"], idx[yr], 0, idx[m]] = (
                ch[m][j] / total_ch[j]
            )
            dm_modal.array[idx["Vaud"], idx[yr], 0, idx[m]] = vd[m][j] / total_vd[j]

    # Extra countries duplicate Switzerland
    for country in all_countries:
        if country not in ("Switzerland", "Vaud"):
            dm_modal.array[idx[country], :, :, :] = dm_modal.array[
                idx["Switzerland"], :, :, :
            ]

    # --- DataMatrix: freight_tkm -----------------------------------------
    dm_tkm = DataMatrix(
        col_labels={
            "Country": all_countries,
            "Years": years_ots,
            "Variables": ["tra_freight_tkm-total-demand"],
        },
        units={"tra_freight_tkm-total-demand": "bn-tkm"},
    )
    idx_t = dm_tkm.idx

    dm_tkm.array[idx_t["Switzerland"], :, 0] = total_ch / 1e3
    dm_tkm.array[idx_t["Vaud"], :, 0] = total_vd / 1e3
    for country in all_countries:
        if country not in ("Switzerland", "Vaud"):
            dm_tkm.array[idx_t[country], :, 0] = dm_tkm.array[
                idx_t["Switzerland"], :, 0
            ]

    return {"freight_tkm": dm_tkm, "freight_modal-share": dm_modal}
