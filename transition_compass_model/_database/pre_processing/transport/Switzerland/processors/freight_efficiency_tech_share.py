"""
Swiss freight OTS: vehicle efficiency and technology share for new vehicles.

Variables
---------
tra_freight_vehicle-efficiency_new : MJ/km
    Energy intensity of new vehicles entering the fleet, by mode × technology.

    Road (HDVH, HDVL, HDVM):
        ICE-diesel = fleet-average ≈ EP2050 HGV/LCV fuel energy (PJ) ÷ VKM (Mkm) × 1000.
        HDVH and HDVM are split from the HGV aggregate using fleet-proportion weighting
        and a fixed HDVH/HDVM efficiency ratio (_HDVH_HDVM_EFF_RATIO = 1.38) reflecting
        the heavier payload of articulated vs rigid trucks.
        Other technologies: ICE-diesel × fixed ratio from physics/literature.

    Non-road (rail, IWW, marine, aviation): fixed literature constants.

tra_freight_technology-share_new : %
    Fraction of new vehicles by technology, by mode × technology.

    Road 1990–2017: ICE-diesel = 100 % (diesel dominance).
    Road 2018–2023: from EP2050 Neuzulassungen (new registrations by segment),
        mapping segments → model technologies (see _SEG_TO_TECH dicts).
    Non-road: fixed literature constants (see _NON_ROAD).

Sources
-------
EP2050 Verkehrssektor ZERO-Basis:
  Sheet 04 Energieverbrauch Strasse: HGV/LCV fuel energy (PJ), 1990-2060.
  Sheet 03 Fahrleistung: HGV/LCV total VKM (Mio. km), 1990-2060.
  Sheet 02 Flottenbestand: fleet counts by segment (ZERO-Basis), 1990-2060.
  Sheet 05 Neuzulassungen Strasse: new registrations by segment (fractions), 2018-2060.
  # https://www.bfe.admin.ch/bfe/en/home/policy/energy-perspectives-2050-plus.html
  # file: EP2050+_Detailergebnisse 2020-2060_Verkehrssektor_alle Szenarien_2022-04-12.xlsx

Non-road efficiency constants:
  Rail CEV  0.117 MJ/km — SBB freight electric, consistent with IEA/UIC rail energy intensity.
  Rail ICE  0.316 MJ/km — Swiss diesel freight locomotive, EEA emission inventory reference.
  IWW       0.429 MJ/km — Rhine barge literature value (CCNR/IWT baseline).
  Marine    0.160 MJ/km — placeholder (Switzerland landlocked; tkm = 0 in model).
  Aviation 19.077 MJ/km — cargo aircraft, consistent with IATA fuel efficiency data.
"""

import os

import numpy as np
import openpyxl
import pandas as pd

from transition_compass_model.model.common.data_matrix_class import DataMatrix

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
# HBEFA 4.2: articulated trucks (TT/AT) consume ~38 % more fuel/km than rigid trucks
# at the same payload class due to higher gross vehicle mass and aerodynamic drag.
# https://www.hbefa.net/
_HDVH_HDVM_EFF_RATIO = 1.38  # HDVH ICE-diesel efficiency / HDVM ICE-diesel efficiency

# Efficiency ratios relative to ICE-diesel (MJ/km basis, road modes).
# ICE-gasoline +12 %: HBEFA 4.2 / COPERT petrol vs diesel heavy-duty vehicles.
#   https://www.hbefa.net/  https://www.emisia.com/utilities/copert/
# ICE-gas (CNG/LNG) ≈ same as diesel in MJ/km; IEA Future of Trucks (2017).
#   https://www.iea.org/reports/the-future-of-trucks
# BEV/CEV 30 % of diesel: electric drivetrain ~70 % more efficient; same IEA source.
# FCEV 33 %: ~10 % less efficient than BEV due to H2 compression/liquefaction losses.
# PHEV-diesel 55 %, PHEV-gasoline 62 %: 50 % electric driving fraction assumed.
#   ICCT HDV technology roadmap: https://theicct.org/publication/hdv-techno-roadmap-may21/
_ROAD_EFF_RATIOS = {
    "ICE-diesel": 1.00,
    "ICE-gasoline": 1.12,
    "ICE-gas": 0.99,
    "BEV": 0.30,
    "CEV": 0.30,
    "FCEV": 0.33,
    "PHEV-diesel": 0.55,
    "PHEV-gasoline": 0.62,
}

# Non-road efficiency (MJ/km) and baseline tech shares.
# Rail CEV 0.117 MJ/km: SBB annual energy report, electric freight traction.
#   https://company.sbb.ch/en/the-company/reports-and-key-figures/annual-report.html
# Rail ICE-diesel 0.316 MJ/km: EEA EMEP/EEA air pollutant emission guidebook 2023.
#   https://www.eea.europa.eu/publications/emep-eea-guidebook-2023
# IWW 0.429 MJ/km: CCNR inland navigation market observation report (Rhine barge).
#   https://www.ccr-zkr.org/
# Marine 0.160 MJ/km: placeholder (Switzerland landlocked, model sets tkm = 0).
# Aviation 19.077 MJ/km: IATA fuel efficiency data, cargo aircraft baseline.
#   https://www.iata.org/en/programs/ops-infra/fuel/
_NON_ROAD = {
    "rail": {
        "eff": {"CEV": 0.117, "ICE-diesel": 0.316},
        "share": {"CEV": 0.90, "ICE-diesel": 0.10},
    },
    "IWW": {
        "eff": {"ICE": 0.429, "BEV": 0.429, "FCEV": 0.429},
        "share": {"ICE": 1.0},
    },
    "marine": {
        "eff": {"ICE": 0.160, "BEV": 0.160, "FCEV": 0.160},
        "share": {"ICE": 1.0},
    },
    "aviation": {
        "eff": {"kerosene": 19.077, "BEV": 19.077},
        "share": {"kerosene": 1.0},
    },
}

# EP2050 Neuzulassungen segment → model technology, for HDVH (TT/AT)
_HDVH_SEG = {
    "TT/AT <=7,5t": "ICE-diesel",
    "TT/AT >7,5-14t": "ICE-diesel",
    "TT/AT >14-20t": "ICE-diesel",
    "TT/AT >20-28t": "ICE-diesel",
    "TT/AT >28-34t": "ICE-diesel",
    "TT/AT >34-40t": "ICE-diesel",
    "TT/AT >40-50t": "ICE-diesel",
    "TT/AT >50-60t": "ICE-diesel",
    "TT/AT >60t": "ICE-diesel",
    "TT/AT CNG": "ICE-gas",
    "TT/AT LNG": "ICE-gas",
    "TT/AT BEV": "BEV",
    "TT/AT PHEV": "PHEV-diesel",
    "TT/AT FCEV": "FCEV",
    "TT/AT <28t EE": "CEV",
    "TT/AT 28-34t EE": "CEV",
}

# EP2050 Neuzulassungen segment → model technology, for HDVM (RigidTruck)
_HDVM_SEG = {
    "RT petrol": "ICE-gasoline",
    "RigidTruck <7,5t": "ICE-diesel",
    "RigidTruck 7,5-12t": "ICE-diesel",
    "RigidTruck >12-14t": "ICE-diesel",
    "RigidTruck >14-20t": "ICE-diesel",
    "RigidTruck >20-26t": "ICE-diesel",
    "RigidTruck >26-28t": "ICE-diesel",
    "RigidTruck >28-32t": "ICE-diesel",
    "RigidTruck >32t": "ICE-diesel",
    "RigidTruck CNG ≤7,5t": "ICE-gas",
    "RigidTruck CNG >7,5-12t": "ICE-gas",
    "RigidTruck CNG >12t": "ICE-gas",
    "RigidTruck LNG <=7.5t": "ICE-gas",
    "RigidTruck LNG  >7.5-12t": "ICE-gas",
    "RigidTruck LNG >12t": "ICE-gas",
    "RigidTruck BEV ≤7.5t": "BEV",
    "RigidTruck BEV >7.5-12t": "BEV",
    "RigidTruck BEV >12t": "BEV",
    "RigidTruck PHEV <=7,5t": "PHEV-diesel",
    "RigidTruck PHEV >7,5-12t": "PHEV-diesel",
    "RigidTruck PHEV >12t": "PHEV-diesel",
    "RigidTruck FCEV <=7,5t": "FCEV",
    "RigidTruck FCEV >7,5-12t": "FCEV",
    "RigidTruck FCEV >12t": "FCEV",
    "RigidTruck <=7,5t EE": "CEV",
    "RigidTruck >7,5-12t EE": "CEV",
}

# EP2050 Neuzulassungen segment → model technology, for HDVL (LCV)
_HDVL_SEG = {
    "LCV petrol M+N1-I": "ICE-gasoline",
    "LCV petrol N1-II": "ICE-gasoline",
    "LCV petrol N1-III": "ICE-gasoline",
    "LCV diesel M+N1-I": "ICE-diesel",
    "LCV diesel N1-II": "ICE-diesel",
    "LCV diesel N1-III": "ICE-diesel",
    "LCV CNG/petrol M+N1-I": "ICE-gas",
    "LCV CNG/petrol N1-II": "ICE-gas",
    "LCV CNG/petrol N1-III": "ICE-gas",
    "LCV BEV M+N1-I": "BEV",
    "LCV BEV N1-II": "BEV",
    "LCV BEV N1-III": "BEV",
    "LCV PHEV diesel N1-III": "PHEV-diesel",
    "LCV FuelCell N1-III": "FCEV",
}

_ALL_TECHS = [
    "BEV",
    "CEV",
    "FCEV",
    "H2",
    "ICE",
    "ICE-diesel",
    "ICE-gas",
    "ICE-gasoline",
    "PHEV-diesel",
    "PHEV-gasoline",
    "kerosene",
]
_ROAD_MODES = ["HDVH", "HDVL", "HDVM"]
_ALL_MODES = ["HDVH", "HDVL", "HDVM", "IWW", "aviation", "marine", "rail"]

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR = os.path.join(_THIS_DIR, "../data")
_EP2050_PATH = os.path.join(
    _DATA_DIR,
    "EP2050_sectors/EP2050+_Szenarienergebnisse_Details_Nachfragesektoren/"
    "EP2050+_Detailergebnisse 2020-2060_Verkehrssektor_alle Szenarien_2022-04-12.xlsx",
)


# ---------------------------------------------------------------------------
# EP2050 readers
# ---------------------------------------------------------------------------
def _read_ep2050_energy() -> dict:
    """Return {veh_cat: {fuel: pd.Series(year→PJ)}} for HGV and LCV."""
    wb = openpyxl.load_workbook(_EP2050_PATH, read_only=True, data_only=True)
    ws = wb["04 Energieverbrauch Strasse"]
    rows = list(ws.iter_rows(min_row=20, max_row=200, values_only=True))
    header = rows[0]
    years_ep = [int(v) for v in header[4:] if isinstance(v, (int, float))]
    n = len(years_ep)

    result = {"HGV": {}, "LCV": {}}
    for r in rows[1:]:
        if r[0] is not None and isinstance(r[0], str) and r[0].startswith("ZERO"):
            break
        cat = r[1]
        if cat not in ("HGV", "LCV"):
            continue
        fuel = r[2]
        vals = [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n]]
        result[cat][fuel] = pd.Series(dict(zip(years_ep, vals)))
    return result


def _read_ep2050_vkm() -> dict:
    """Return {cat: pd.Series(year→million_vkm)} for HGV and LCV."""
    wb = openpyxl.load_workbook(_EP2050_PATH, read_only=True, data_only=True)
    ws = wb["03 Fahrleistung"]
    rows = list(ws.iter_rows(min_row=20, max_row=300, values_only=True))
    header = rows[0]
    years_ep = [int(v) for v in header[4:] if isinstance(v, (int, float))]
    n = len(years_ep)

    lcv = np.zeros(n)
    hgv = np.zeros(n)
    in_lcv = in_hgv = done_lcv = done_hgv = False
    for r in rows[1:]:
        if r[1] == "LCV" and not done_lcv:
            lcv += np.array(
                [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n]]
            )
            in_lcv = True
        elif in_lcv and r[1] != "LCV":
            done_lcv = True
        if r[1] == "HGV" and not done_hgv:
            hgv += np.array(
                [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n]]
            )
            in_hgv = True
        elif in_hgv and r[1] != "HGV":
            done_hgv = True
        if done_lcv and done_hgv:
            break
    return {
        "HGV": pd.Series(dict(zip(years_ep, hgv))),
        "LCV": pd.Series(dict(zip(years_ep, lcv))),
    }


def _read_ep2050_fleet_frac(years_list: list) -> np.ndarray:
    """Return HDVH fraction of HGV fleet (array aligned to years_list)."""
    wb = openpyxl.load_workbook(_EP2050_PATH, read_only=True, data_only=True)
    ws = wb["02 Flottenbestand"]
    rows = list(ws.iter_rows(min_row=20, max_row=300, values_only=True))
    header = rows[0]
    years_ep = [int(v) for v in header[4:] if isinstance(v, (int, float))]
    n = len(years_ep)

    hdvh = np.zeros(n)
    hdvm = np.zeros(n)
    for r in rows[1:]:
        if r[0] is not None and isinstance(r[0], str) and r[0].startswith("ZERO"):
            break
        if r[1] is None:
            continue
        vals = np.array(
            [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n]]
        )
        if r[1].startswith("TT/AT"):
            hdvh += vals
        elif r[1].startswith("RigidTruck"):
            hdvm += vals

    total = hdvh + hdvm
    frac = np.where(total > 0, hdvh / total, 0.25)
    raw = pd.Series(dict(zip(years_ep, frac)))
    filled = raw.reindex(years_list, fill_value=np.nan).astype(float)
    return filled.interpolate(method="index").bfill().ffill().values


def _read_ep2050_tech_shares_road(years_ots: list) -> dict:
    """Return {mode: {tech: np.ndarray(len(years_ots))}} from Neuzulassungen 2018+.

    For years before 2018: ICE-diesel = 1.0 for all road modes.
    """
    wb = openpyxl.load_workbook(_EP2050_PATH, read_only=True, data_only=True)
    ws = wb["05 Neuzulassungen Strasse"]
    rows = list(ws.iter_rows(min_row=20, max_row=200, values_only=True))
    header = rows[0]
    years_ep = [int(v) for v in header[4:] if isinstance(v, (int, float))]
    n_ep = len(years_ep)

    # Accumulate fractions (of HGV or LCV total) by model tech
    hgv_hdvh = {t: np.zeros(n_ep) for t in set(_HDVH_SEG.values())}
    hgv_hdvm = {t: np.zeros(n_ep) for t in set(_HDVM_SEG.values())}
    lcv_hdvl = {t: np.zeros(n_ep) for t in set(_HDVL_SEG.values())}

    for r in rows[1:]:
        if r[0] is not None and isinstance(r[0], str) and r[0].startswith("ZERO"):
            break
        cat, seg = r[1], r[2]
        if cat == "HGV" and seg != "Total" and seg in _HDVH_SEG:
            tech = _HDVH_SEG[seg]
            hgv_hdvh[tech] += np.array(
                [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n_ep]]
            )
        elif cat == "HGV" and seg != "Total" and seg in _HDVM_SEG:
            tech = _HDVM_SEG[seg]
            hgv_hdvm[tech] += np.array(
                [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n_ep]]
            )
        elif cat == "LCV" and seg != "Total" and seg in _HDVL_SEG:
            tech = _HDVL_SEG[seg]
            lcv_hdvl[tech] += np.array(
                [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n_ep]]
            )

    def _normalize(tech_dict):
        total = sum(tech_dict.values())
        if np.all(total == 0):
            return {t: np.zeros(n_ep) for t in tech_dict}
        return {t: np.where(total > 0, v / total, 0.0) for t, v in tech_dict.items()}

    hdvh_norm = _normalize(hgv_hdvh)
    hdvm_norm = _normalize(hgv_hdvm)
    hdvl_norm = _normalize(lcv_hdvl)

    def _fill_to_ots(norm_dict, seg_map):
        """For each tech, produce an array over years_ots.
        EP2050 covers 2018+; before 2018 → ICE-diesel = 1.0.
        """
        result = {}
        ep_ser = {t: pd.Series(dict(zip(years_ep, v))) for t, v in norm_dict.items()}
        for tech in set(seg_map.values()):
            raw = ep_ser.get(tech, pd.Series(dtype=float))
            aligned = raw.reindex(years_ots, fill_value=np.nan).astype(float)
            # Fill pre-2018 with 0 (ICE-diesel gets the residual below)
            aligned.loc[aligned.index < 2018] = 0.0
            # Forward-fill post-2023 gaps
            aligned = aligned.interpolate(method="index").ffill()
            result[tech] = aligned.values

        # ICE-diesel = 1 − everything else for pre-2018
        ice_diesel = result.get("ICE-diesel", np.zeros(len(years_ots)))
        other_sum = sum(v for t, v in result.items() if t != "ICE-diesel")
        for i, yr in enumerate(years_ots):
            if yr < 2018:
                for t in result:
                    result[t][i] = 1.0 if t == "ICE-diesel" else 0.0
            else:
                result["ICE-diesel"][i] = max(
                    0.0, ice_diesel[i] + (1.0 - other_sum[i] - ice_diesel[i])
                )
        return result

    return {
        "HDVH": _fill_to_ots(hdvh_norm, _HDVH_SEG),
        "HDVM": _fill_to_ots(hdvm_norm, _HDVM_SEG),
        "HDVL": _fill_to_ots(hdvl_norm, _HDVL_SEG),
    }


def _fill(raw: pd.Series, years: list) -> np.ndarray:
    s = raw.reindex(years, fill_value=np.nan).astype(float)
    s = s.interpolate(method="index")
    return s.bfill().ffill().values


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def run(years_ots: list, country_list: list) -> dict:
    """Build freight_vehicle-efficiency_new and freight_technology-share_new DataMatrices.

    Parameters
    ----------
    years_ots : list of int
    country_list : list of str

    Returns
    -------
    dict with keys 'freight_vehicle-efficiency_new' and 'freight_technology-share_new',
    each a DataMatrix with Countries, Years=years_ots, Variables, Cat1=modes, Cat2=techs.
    """
    all_countries = ["Switzerland", "Vaud"] + [
        c for c in country_list if c not in ("Switzerland", "Vaud")
    ]

    # --- Load EP2050 data ---
    ep_energy = _read_ep2050_energy()
    ep_vkm = _read_ep2050_vkm()
    frac_hdvh = _read_ep2050_fleet_frac(years_ots)
    road_shares = _read_ep2050_tech_shares_road(years_ots)

    # --- ICE-diesel fleet-avg efficiency per road category (MJ/km) ---
    # efficiency(MJ/km) = energy(PJ) × 1000 / VKM(Mkm)
    hgv_diesel_energy = _fill(
        ep_energy["HGV"].get("diesel", pd.Series(dtype=float)), years_ots
    )
    hgv_vkm = _fill(ep_vkm["HGV"], years_ots)
    lcv_diesel_energy = _fill(
        ep_energy["LCV"].get("diesel", pd.Series(dtype=float)), years_ots
    )
    lcv_vkm = _fill(ep_vkm["LCV"], years_ots)

    hgv_avg_eff = np.where(hgv_vkm > 0, hgv_diesel_energy * 1000.0 / hgv_vkm, np.nan)
    # Split HDVH / HDVM using size ratio
    R = _HDVH_HDVM_EFF_RATIO
    denom = frac_hdvh * R + (1 - frac_hdvh)
    hdvm_eff = np.where(denom > 0, hgv_avg_eff / denom, np.nan)
    hdvh_eff = R * hdvm_eff
    hdvl_eff = np.where(lcv_vkm > 0, lcv_diesel_energy * 1000.0 / lcv_vkm, np.nan)

    road_base_eff = {"HDVH": hdvh_eff, "HDVM": hdvm_eff, "HDVL": hdvl_eff}

    # --- Build DataMatrices ---
    def _make_dm(varname, unit):
        return DataMatrix(
            col_labels={
                "Country": all_countries,
                "Years": years_ots,
                "Variables": [varname],
                "Categories1": _ALL_MODES,
                "Categories2": _ALL_TECHS,
            },
            units={varname: unit},
        )

    dm_eff = _make_dm("tra_freight_vehicle-efficiency_new", "MJ/km")
    dm_share = _make_dm("tra_freight_technology-share_new", "%")
    idx_e = dm_eff.idx
    idx_s = dm_share.idx
    ch = "Switzerland"

    # --- Road modes: efficiency ---
    for mode in _ROAD_MODES:
        base_eff = road_base_eff[mode]
        for tech, ratio in _ROAD_EFF_RATIOS.items():
            if tech not in idx_e:
                continue
            dm_eff.array[idx_e[ch], :, 0, idx_e[mode], idx_e[tech]] = base_eff * ratio

    # --- Road modes: tech share ---
    for mode in _ROAD_MODES:
        shares_by_tech = road_shares[mode]
        for tech, vals in shares_by_tech.items():
            if tech not in idx_s:
                continue
            dm_share.array[idx_s[ch], :, 0, idx_s[mode], idx_s[tech]] = vals

    # --- Non-road modes: efficiency and tech share ---
    for mode, data in _NON_ROAD.items():
        if mode not in idx_e:
            continue
        for tech, val in data["eff"].items():
            if tech in idx_e:
                dm_eff.array[idx_e[ch], :, 0, idx_e[mode], idx_e[tech]] = val
        for tech, val in data["share"].items():
            if tech in idx_s:
                dm_share.array[idx_s[ch], :, 0, idx_s[mode], idx_s[tech]] = val

    # --- Copy Switzerland to all other countries ---
    for country in all_countries:
        if country != ch:
            dm_eff.array[idx_e[country], :, :, :, :] = dm_eff.array[
                idx_e[ch], :, :, :, :
            ]
            dm_share.array[idx_s[country], :, :, :, :] = dm_share.array[
                idx_s[ch], :, :, :, :
            ]

    return {
        "freight_vehicle-efficiency_new": dm_eff,
        "freight_technology-share_new": dm_share,
    }
