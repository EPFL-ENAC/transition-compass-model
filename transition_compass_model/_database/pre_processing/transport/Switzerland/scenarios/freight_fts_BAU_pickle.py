"""
Swiss freight FTS: all-scenarios freight FTS generation.

Level 1 (BAU)   : flat continuation / linear OTS trend.
Level 4 (ambitious): EP2050 ZERO-B where clean data is available.
Levels 2-3      : linear interpolation (midpoints) between 1 and 4.

EP2050 ZERO-B used for road modes (HDVH, HDVL, HDVM) only:
  freight_technology-share_new – sheet 05 Neuzulassungen Strasse
  freight_utilization-rate     – VKM (sheet 03) / fleet (sheet 02)

All other keys and non-road modes use level 1 for all 4 levels.
# TODO levels 2-4:
#   freight_tkm              – Swiss ARE demand scenario projections
#   freight_modal-share      – modal shift to rail (BAV/NEAT targets)
#   freight_vehicle-efficiency_new – per-tech efficiency improvements (EP2050 ZERO-B)
#   non-road modes (rail, IWW, marine, aviation) in freight_technology-share_new

Sources
-------
EP2050 Verkehrssektor:
  # https://www.bfe.admin.ch/bfe/en/home/policy/energy-perspectives-2050-plus.html
  # file: EP2050+_Detailergebnisse 2020-2060_Verkehrssektor_alle Szenarien_2022-04-12.xlsx
"""

import os

import numpy as np
import openpyxl
import pandas as pd
from processors.freight_efficiency_tech_share import (
    _EP2050_PATH,
    _HDVH_SEG,
    _HDVL_SEG,
    _HDVM_SEG,
)

from transition_compass_model.model.common.auxiliary_functions import (
    create_years_list,
    linear_fitting,
    my_pickle_dump,
    sort_pickle,
)

_ROAD_MODES = ["HDVH", "HDVL", "HDVM"]


# ---------------------------------------------------------------------------
# EP2050 ZERO-B helpers
# ---------------------------------------------------------------------------
def _find_zerob_rows(ws):
    """Return (header_row, data_rows) for the ZERO-B sub-table in a sheet.

    Each EP2050 sheet stores three scenario sub-tables (ZERO-Basis, ZERO-A,
    ZERO-B) with an identical 5-row preamble: scenario label, blank, 'zurück',
    table title, description, then the column header.
    """
    all_rows = list(ws.iter_rows(min_row=1, max_row=500, values_only=True))
    start = None
    for i, r in enumerate(all_rows):
        if r[0] is not None and str(r[0]) == "ZERO-B":
            start = i + 5  # header is always 5 rows after the scenario label
            break
    if start is None:
        raise ValueError("ZERO-B not found in sheet")
    header = all_rows[start]
    data_rows = []
    for r in all_rows[start + 1 :]:
        if r[0] is not None and isinstance(r[0], str) and r[0].startswith("ZERO"):
            break
        data_rows.append(r)
    return header, data_rows


def _zerob_tech_shares_road(years_fts):
    """Read ZERO-B new-vehicle tech shares for HDVH, HDVL, HDVM.

    Returns {mode: {tech: np.ndarray(len(years_fts))}}.
    """
    wb = openpyxl.load_workbook(_EP2050_PATH, read_only=True, data_only=True)
    header, data_rows = _find_zerob_rows(wb["05 Neuzulassungen Strasse"])
    years_ep = [int(v) for v in header[4:] if isinstance(v, (int, float))]
    n_ep = len(years_ep)

    hgv_hdvh = {t: np.zeros(n_ep) for t in set(_HDVH_SEG.values())}
    hgv_hdvm = {t: np.zeros(n_ep) for t in set(_HDVM_SEG.values())}
    lcv_hdvl = {t: np.zeros(n_ep) for t in set(_HDVL_SEG.values())}

    for r in data_rows:
        cat, seg = r[1], r[2]
        if not cat or not seg or seg == "Total":
            continue
        vals = np.array(
            [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n_ep]]
        )
        if cat == "HGV" and seg in _HDVH_SEG:
            hgv_hdvh[_HDVH_SEG[seg]] += vals
        elif cat == "HGV" and seg in _HDVM_SEG:
            hgv_hdvm[_HDVM_SEG[seg]] += vals
        elif cat == "LCV" and seg in _HDVL_SEG:
            lcv_hdvl[_HDVL_SEG[seg]] += vals

    def _normalize(tech_dict):
        total = sum(tech_dict.values())
        if np.all(total == 0):
            return {t: np.zeros(n_ep) for t in tech_dict}
        return {t: np.where(total > 0, v / total, 0.0) for t, v in tech_dict.items()}

    def _align(norm_dict):
        result = {}
        for tech, arr in norm_dict.items():
            raw = pd.Series(dict(zip(years_ep, arr)))
            aligned = raw.reindex(years_fts, fill_value=np.nan).astype(float)
            result[tech] = aligned.interpolate(method="index").ffill().bfill().values
        # Renormalise after reindexing to ensure exact sum-to-1
        totals = sum(result.values())
        return {t: np.where(totals > 0, v / totals, 0.0) for t, v in result.items()}

    return {
        "HDVH": _align(_normalize(hgv_hdvh)),
        "HDVM": _align(_normalize(hgv_hdvm)),
        "HDVL": _align(_normalize(lcv_hdvl)),
    }


def _zerob_utilization_road(years_fts):
    """Read ZERO-B utilisation rate (vkm/veh/yr) for HDVH, HDVL, HDVM.

    Returns {mode: np.ndarray(len(years_fts))}.
    """
    wb = openpyxl.load_workbook(_EP2050_PATH, read_only=True, data_only=True)

    # VKM (sheet 03)
    header_v, rows_v = _find_zerob_rows(wb["03 Fahrleistung"])
    years_ep = [int(v) for v in header_v[4:] if isinstance(v, (int, float))]
    n = len(years_ep)
    lcv = np.zeros(n)
    hgv = np.zeros(n)
    in_lcv = in_hgv = done_lcv = done_hgv = False
    for r in rows_v:
        if r[1] == "LCV" and not done_lcv:
            lcv += np.array(
                [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n]]
            )
            in_lcv = True
        elif in_lcv and r[1] != "LCV":
            done_lcv = True
        if r[1] == "HGV" and not done_hgv:
            hgv += np.array(
                [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n]]
            )
            in_hgv = True
        elif in_hgv and r[1] != "HGV":
            done_hgv = True
        if done_lcv and done_hgv:
            break

    # Fleet (sheet 02)
    header_f, rows_f = _find_zerob_rows(wb["02 Flottenbestand"])
    hdvh_f = np.zeros(n)
    hdvm_f = np.zeros(n)
    hdvl_f = np.zeros(n)
    for r in rows_f:
        if r[1] is None:
            continue
        vals = np.array(
            [v if isinstance(v, (int, float)) else 0.0 for v in r[4 : 4 + n]]
        )
        if r[1].startswith("TT/AT"):
            hdvh_f += vals
        elif r[1].startswith("RigidTruck"):
            hdvm_f += vals
        elif r[1].startswith("LCV"):
            hdvl_f += vals

    total_hgv = hdvh_f + hdvm_f
    frac = np.where(total_hgv > 0, hdvh_f / total_hgv, 0.25)
    hdvh_vkm = hgv * frac
    hdvm_vkm = hgv * (1 - frac)

    def _align(mkm, fleet):
        ur = np.where(fleet > 0, mkm * 1e6 / fleet, np.nan)
        raw = pd.Series(dict(zip(years_ep, ur)))
        return (
            raw.reindex(years_fts, fill_value=np.nan)
            .astype(float)
            .interpolate(method="index")
            .ffill()
            .bfill()
            .values
        )

    return {
        "HDVH": _align(hdvh_vkm, hdvh_f),
        "HDVM": _align(hdvm_vkm, hdvm_f),
        "HDVL": _align(lcv, hdvl_f),
    }


# ---------------------------------------------------------------------------
# Midpoint helper
# ---------------------------------------------------------------------------
def _midpoint(dm1, dm4, frac):
    """Linear interpolation: dm1 + frac * (dm4 - dm1)."""
    dm_out = dm1.copy()
    dm_out.array = dm1.array + frac * (dm4.array - dm1.array)
    return dm_out


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def run(DM_transport, country_list, years_ots, years_fts):
    """Build freight FTS DataMatrices for all 4 lever levels and save to pickle.

    Parameters
    ----------
    DM_transport : dict
        Full transport DM dict (after passenger BAU run).
    country_list : list of str
    years_ots : list of int
    years_fts : list of int
    """
    all_countries = ["Switzerland", "Vaud"] + [
        c for c in country_list if c not in ("Switzerland", "Vaud")
    ]
    ch = "Switzerland"
    DM_fts = {"fts": {}}

    # ------------------------------------------------------------------
    # freight_tkm : linear OTS trend, all 4 levels identical
    # TODO levels 2-4: Swiss ARE freight demand scenario projections
    # ------------------------------------------------------------------
    dm_tkm = DM_transport["ots"]["freight_tkm"].copy()
    linear_fitting(dm_tkm, years_fts, based_on=create_years_list(2010, 2019, 1))
    dm_tkm_fts = dm_tkm.filter({"Years": years_fts})
    DM_fts["fts"]["freight_tkm"] = {lev: dm_tkm_fts.copy() for lev in range(1, 5)}

    # ------------------------------------------------------------------
    # freight_modal-share : flat continuation, all 4 levels identical
    # TODO levels 2-4: modal shift to rail (BAV/NEAT targets)
    # ------------------------------------------------------------------
    dm_ms = DM_transport["ots"]["freight_modal-share"].copy()
    dm_ms.add(np.nan, dim="Years", col_label=years_fts, dummy=True)
    dm_ms.fill_nans("Years")
    dm_ms_fts = dm_ms.filter({"Years": years_fts})
    DM_fts["fts"]["freight_modal-share"] = {
        lev: dm_ms_fts.copy() for lev in range(1, 5)
    }

    # ------------------------------------------------------------------
    # freight_vehicle-efficiency_new : flat continuation, all 4 levels identical
    # TODO levels 2-4: per-tech efficiency improvements (EP2050 ZERO-B)
    # ------------------------------------------------------------------
    dm_eff = DM_transport["ots"]["freight_vehicle-efficiency_new"].copy()
    dm_eff.add(np.nan, dim="Years", col_label=years_fts, dummy=True)
    dm_eff.fill_nans("Years")
    dm_eff_fts = dm_eff.filter({"Years": years_fts})
    DM_fts["fts"]["freight_vehicle-efficiency_new"] = {
        lev: dm_eff_fts.copy() for lev in range(1, 5)
    }

    # ------------------------------------------------------------------
    # freight_technology-share_new
    # Level 1  : flat continuation of OTS 2023
    # Level 4  : EP2050 ZERO-B (road modes); non-road = level 1
    # Levels 2-3: midpoints
    # TODO non-road modes (rail, IWW, marine, aviation) levels 2-4
    # ------------------------------------------------------------------
    dm_ts_ots = DM_transport["ots"]["freight_technology-share_new"]

    # Level 1 (flat)
    dm_ts_lev1 = dm_ts_ots.copy()
    dm_ts_lev1.add(np.nan, dim="Years", col_label=years_fts, dummy=True)
    dm_ts_lev1.fill_nans("Years")
    dm_ts_lev1 = dm_ts_lev1.filter({"Years": years_fts})

    # Level 4: override road modes with EP2050 ZERO-B
    zerob_shares = _zerob_tech_shares_road(years_fts)
    dm_ts_lev4 = dm_ts_lev1.copy()
    idx = dm_ts_lev4.idx
    for mode in _ROAD_MODES:
        # Zero out this mode first, then fill with ZERO-B values
        dm_ts_lev4.array[idx[ch], :, 0, idx[mode], :] = 0.0
        for tech, vals in zerob_shares[mode].items():
            if tech in idx:
                dm_ts_lev4.array[idx[ch], :, 0, idx[mode], idx[tech]] = vals
        for country in all_countries:
            if country != ch:
                dm_ts_lev4.array[idx[country], :, 0, idx[mode], :] = dm_ts_lev4.array[
                    idx[ch], :, 0, idx[mode], :
                ]

    dm_ts_lev2 = _midpoint(dm_ts_lev1, dm_ts_lev4, 1 / 3)
    dm_ts_lev3 = _midpoint(dm_ts_lev1, dm_ts_lev4, 2 / 3)
    DM_fts["fts"]["freight_technology-share_new"] = {
        1: dm_ts_lev1,
        2: dm_ts_lev2,
        3: dm_ts_lev3,
        4: dm_ts_lev4,
    }

    # ------------------------------------------------------------------
    # freight_utilization-rate
    # Level 1  : flat continuation of OTS 2023 (load-factor + utilisation-rate)
    # Level 4  : EP2050 ZERO-B utilisation-rate for road; load-factor = level 1
    # Levels 2-3: midpoints
    # TODO load-factor levels 2-4 (no EP2050 scenario data available)
    # ------------------------------------------------------------------
    dm_ur_ots = DM_transport["ots"]["freight_utilization-rate"]

    # Level 1 (flat)
    dm_ur_lev1 = dm_ur_ots.copy()
    dm_ur_lev1.add(np.nan, dim="Years", col_label=years_fts, dummy=True)
    dm_ur_lev1.fill_nans("Years")
    dm_ur_lev1 = dm_ur_lev1.filter({"Years": years_fts})

    # Level 4: override utilisation-rate with ZERO-B; load-factor unchanged
    zerob_ur = _zerob_utilization_road(years_fts)
    dm_ur_lev4 = dm_ur_lev1.copy()
    idx = dm_ur_lev4.idx
    var_ur = "tra_freight_utilisation-rate"
    for mode in _ROAD_MODES:
        dm_ur_lev4.array[idx[ch], :, idx[var_ur], idx[mode]] = zerob_ur[mode]
        for country in all_countries:
            if country != ch:
                dm_ur_lev4.array[idx[country], :, idx[var_ur], idx[mode]] = zerob_ur[
                    mode
                ]

    dm_ur_lev2 = _midpoint(dm_ur_lev1, dm_ur_lev4, 1 / 3)
    dm_ur_lev3 = _midpoint(dm_ur_lev1, dm_ur_lev4, 2 / 3)
    DM_fts["fts"]["freight_utilization-rate"] = {
        1: dm_ur_lev1,
        2: dm_ur_lev2,
        3: dm_ur_lev3,
        4: dm_ur_lev4,
    }

    # Save
    this_dir = os.path.dirname(os.path.abspath(__file__))
    pickle_file = os.path.join(this_dir, "../../../../data/datamatrix/transport.pickle")
    my_pickle_dump(DM_new=DM_fts, local_pickle_file=pickle_file)
    sort_pickle(pickle_file)

    return DM_transport
